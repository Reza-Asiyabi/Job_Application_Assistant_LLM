#!/usr/bin/env python3
"""
import_tracker.py
Convert a CSV or XLSX application-tracking spreadsheet into applications.json.

Usage
-----
  python import_tracker.py                          # interactive: prompts for file
  python import_tracker.py my_apps.csv              # direct CSV import
  python import_tracker.py "What I've done.xlsx"    # direct XLSX import
  python import_tracker.py my_apps.csv --replace    # overwrite instead of merging

XLSX support requires openpyxl:
  pip install openpyxl
"""

from __future__ import annotations

import csv
import json
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

APPLICATIONS_FILE = "applications.json"

STATUSES = [
    "Watching", "Applied", "Phone Screen", "Interview",
    "Final Round", "Offer", "Accepted", "Rejected", "Withdrawn",
]

# ─────────────────────────────────────────────────────────────────────────────
# Column auto-detection
# Each field maps to a list of substrings that, when found anywhere in a column
# header (case-insensitive), indicate that column should map to that field.
# More specific patterns come first so they win over generic ones.
# ─────────────────────────────────────────────────────────────────────────────
FIELD_PATTERNS: dict[str, list[str]] = {
    "company":      ["company", "employer", "organisation", "organization", "firm", "corp"],
    "position":     ["position", "role", "job title", "title", "job", "post", "opening", "vacancy"],
    "location":     ["location", "place", "city", "office", "country", "where", "region"],
    "status":       ["status", "outcome", "result", "response", "stage", "progress", "state"],
    "salary_jd":    ["salary (jd)", "salary jd", "advertised", "offered salary", "jd salary",
                     "package", "pay", "compensation", "comp", "salary offered"],
    "salary_req":   ["salary req", "requested", "my salary", "asking", "desired",
                     "expected salary", "salary expected", "i asked"],
    "date_applied": ["date applied", "applied on", "application date", "sent date",
                     "date sent", "submitted on", "when applied"],
    "source":       ["source", "applied via", "found via", "platform", "channel",
                     "via", "board", "from"],
    "notes":        ["what i've done", "what i have done", "what was done", "notes",
                     "comments", "description", "details", "remarks", "actions",
                     "what", "done", "how applied"],
    # date is intentionally last — it's very generic and should only match
    # when none of the more specific date patterns above matched
    "_date_fallback": ["date"],
}

# ─────────────────────────────────────────────────────────────────────────────
# Status normalisation
# ─────────────────────────────────────────────────────────────────────────────
_STATUS_KEYWORDS: dict[str, list[str]] = {
    "Watching":     ["watching", "interested", "considering", "saved", "bookmarked", "shortlisted"],
    "Applied":      ["applied", "submitted", "sent", "application submitted", "done", "complete"],
    "Phone Screen": ["phone screen", "phone call", "screening call", "recruiter call",
                     "hr call", "screening"],
    "Interview":    ["interview", "interviewing", "technical interview", "onsite",
                     "in-person", "video call", "meeting"],
    "Final Round":  ["final round", "final stage", "final interview", "last round", "last stage"],
    "Offer":        ["offer received", "offer", "offered"],
    "Accepted":     ["accepted", "accepting", "accepted offer", "signed", "yes, accepted"],
    "Rejected":     ["rejected", "rejection", "unsuccessful", "not selected",
                     "not progressed", "no, rejected", "declined by them"],
    "Withdrawn":    ["withdrawn", "withdrew", "cancelled", "dropped", "no longer interested",
                     "declined (me)", "declined offer"],
}


def normalize_status(raw: str) -> str:
    """Map a freeform status string to one of the standard STATUSES values."""
    if not raw or not raw.strip():
        return "Applied"
    raw_lower = raw.lower().strip()
    # Exact match (case-insensitive)
    for s in STATUSES:
        if s.lower() == raw_lower:
            return s
    # Keyword search — longer keywords first to avoid partial false matches
    for status, keywords in _STATUS_KEYWORDS.items():
        for kw in sorted(keywords, key=len, reverse=True):
            if kw in raw_lower:
                return status
    # Single-letter / single-word shortcuts common in personal spreadsheets
    shortcuts = {"r": "Rejected", "a": "Applied", "o": "Offer", "i": "Interview",
                 "w": "Withdrawn", "n": "Rejected", "y": "Accepted"}
    if raw_lower in shortcuts:
        return shortcuts[raw_lower]
    return "Applied"   # safe default


def normalize_date(raw) -> str:
    """Normalise a cell value to YYYY-MM-DD or return raw string."""
    if raw is None:
        return ""
    # openpyxl returns datetime/date objects directly
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s or s.lower() in ("none", "nan", "nat", "-", "n/a"):
        return ""
    # Try common date formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
                "%d.%m.%Y", "%Y/%m/%d", "%d %B %Y", "%d %b %Y",
                "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s   # return as-is if unparseable


def normalize_cell(val) -> str:
    """Convert any cell value to a clean string."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("none", "nan", "nat") else s


# ─────────────────────────────────────────────────────────────────────────────
# Column detection
# ─────────────────────────────────────────────────────────────────────────────

def detect_mapping(headers: list[str]) -> dict[str, int]:
    """Return {field: column_index} by fuzzy header matching."""
    mapping: dict[str, int] = {}
    headers_lower = [h.lower().strip() for h in headers]

    # Process real fields first (skip _date_fallback until end)
    fields_ordered = [f for f in FIELD_PATTERNS if not f.startswith("_")]

    for field in fields_ordered:
        patterns = FIELD_PATTERNS[field]
        for idx, h in enumerate(headers_lower):
            if idx in mapping.values():
                continue   # already claimed
            for pat in patterns:
                if pat in h or h in pat:
                    mapping[field] = idx
                    break
            if field in mapping:
                break

    # Fallback: if date_applied still unmapped, look for any column with "date"
    if "date_applied" not in mapping:
        for idx, h in enumerate(headers_lower):
            if idx in mapping.values():
                continue
            if "date" in h:
                mapping["date_applied"] = idx
                break

    return mapping


# ─────────────────────────────────────────────────────────────────────────────
# File readers
# ─────────────────────────────────────────────────────────────────────────────

def read_csv(path: Path) -> tuple[list[str], list[list]]:
    """Return (headers, data_rows) from a CSV file."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return [], []
    # Find the first non-empty row as the header
    for i, row in enumerate(rows):
        if any(c.strip() for c in row):
            return row, rows[i + 1:]
    return [], []


def read_xlsx(path: Path) -> tuple[list[str], list[list]]:
    """Return (headers, data_rows) from an XLSX file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        print("\n  ERROR: openpyxl is required for XLSX files.")
        print("  Install it with:   pip install openpyxl\n")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)

    # Choose sheet
    sheet_name = wb.sheetnames[0]
    if len(wb.sheetnames) > 1:
        print(f"\n  Sheets in this workbook:")
        for i, name in enumerate(wb.sheetnames, 1):
            print(f"    {i}. {name}")
        choice = input("  Which sheet? (number or name, Enter for first): ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(wb.sheetnames):
            sheet_name = wb.sheetnames[int(choice) - 1]
        elif choice in wb.sheetnames:
            sheet_name = choice

    ws = wb[sheet_name]
    all_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

    # Find first non-empty row as header
    for i, row in enumerate(all_rows):
        if any(v is not None and str(v).strip() not in ("", "None") for v in row):
            headers = [str(v).strip() if v is not None else f"Col{j}"
                       for j, v in enumerate(row)]
            data = [r for r in all_rows[i + 1:]
                    if any(v is not None and str(v).strip() not in ("", "None") for v in r)]
            return headers, data

    return [], []


# ─────────────────────────────────────────────────────────────────────────────
# Interactive mapping confirmation
# ─────────────────────────────────────────────────────────────────────────────

_FIELD_LABELS: dict[str, str] = {
    "company":      "Company name",
    "position":     "Position / role title",
    "location":     "Location",
    "status":       "Status / outcome",
    "salary_jd":    "Salary from JD",
    "salary_req":   "Salary requested",
    "date_applied": "Date applied",
    "source":       "Application source",
    "notes":        "Notes / what was done",
}


def confirm_mapping(headers: list[str], mapping: dict[str, int]) -> dict[str, int]:
    print("\n" + "─" * 64)
    print("  AUTO-DETECTED COLUMN MAPPING")
    print("─" * 64)
    print(f"  {'Field':<22} {'→  Your column':<32} {'Index'}")
    print(f"  {'─'*22} {'─'*32} {'─'*5}")
    for field, label in _FIELD_LABELS.items():
        idx = mapping.get(field)
        col = f"→  {headers[idx]}" if idx is not None else "   — not mapped —"
        print(f"  {label:<22} {col:<32} {idx if idx is not None else '':>5}")

    print()
    print("  All columns in your file:")
    for i, h in enumerate(headers):
        print(f"    {i:>2}: {h}")
    print()

    ans = input("  Accept this mapping? [Y / n]: ").strip().lower()
    if ans not in ("n", "no"):
        return mapping

    # Manual override
    print("\n  Enter the column NUMBER for each field (Enter to keep current / skip):")
    new_mapping: dict[str, int] = {}
    for field, label in _FIELD_LABELS.items():
        current = mapping.get(field)
        hint = f"currently col {current}: {headers[current]}" if current is not None else "not mapped"
        val = input(f"  {label:<24} ({hint}) → ").strip()
        if val.isdigit():
            new_mapping[field] = int(val)
        elif current is not None:
            new_mapping[field] = current
        # else: leave unmapped
    return new_mapping


# ─────────────────────────────────────────────────────────────────────────────
# Row → entry conversion
# ─────────────────────────────────────────────────────────────────────────────

def row_to_entry(row: list, mapping: dict[str, int], seq: int) -> dict:
    """Convert one spreadsheet row to an applications.json entry."""
    def get(field: str) -> str:
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return ""
        return normalize_cell(row[idx])

    date_raw = row[mapping["date_applied"]] if "date_applied" in mapping and mapping["date_applied"] < len(row) else None

    return {
        "id":           f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{seq:04d}",
        "date_applied": normalize_date(date_raw),
        "company":      get("company"),
        "position":     get("position"),
        "location":     get("location"),
        "source":       get("source"),
        "status":       normalize_status(get("status")),
        "salary_jd":    get("salary_jd"),
        "salary_req":   get("salary_req"),
        "notes":        get("notes"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "═" * 64)
    print("  IMPORT TRACKER   CSV / XLSX  →  applications.json")
    print("═" * 64)

    # ── Resolve file path ────────────────────────────────────────────────
    args = sys.argv[1:]
    file_arg = next((a for a in args if not a.startswith("--")), None)
    replace  = "--replace" in args

    if file_arg:
        path = Path(file_arg)
    else:
        raw = input("\n  Path to CSV or XLSX file: ").strip().strip('"').strip("'")
        path = Path(raw)

    if not path.exists():
        print(f"\n  ERROR: File not found: {path}")
        sys.exit(1)

    # ── Read file ────────────────────────────────────────────────────────
    suffix = path.suffix.lower()
    print(f"\n  Reading: {path.name}")
    if suffix == ".csv":
        headers, rows = read_csv(path)
    elif suffix in (".xlsx", ".xls"):
        headers, rows = read_xlsx(path)
    else:
        print(f"  ERROR: Unsupported file type '{suffix}'. Use .csv or .xlsx")
        sys.exit(1)

    # Remove entirely empty rows
    rows = [r for r in rows
            if any(str(v).strip() not in ("", "None", "nan") for v in r if v is not None)]

    if not rows:
        print("  ERROR: No data rows found in file.")
        sys.exit(1)

    print(f"  Found {len(rows):,} data rows across {len(headers)} columns.")

    # ── Detect and confirm column mapping ────────────────────────────────
    mapping = detect_mapping(headers)
    mapping = confirm_mapping(headers, mapping)

    # ── Preview first 3 rows ─────────────────────────────────────────────
    print("\n" + "─" * 64)
    print("  PREVIEW  (first 3 converted rows)")
    print("─" * 64)
    for i, row in enumerate(rows[:3]):
        entry = row_to_entry(row, mapping, i)
        print(f"\n  Row {i + 1}:")
        for k in ["company", "position", "location", "status", "date_applied",
                  "salary_jd", "salary_req", "source"]:
            v = entry.get(k, "")
            if v:
                print(f"    {k:<14} {v[:60]}")
        if entry.get("notes"):
            notes_preview = entry["notes"][:80].replace("\n", " ")
            print(f"    {'notes':<14} {notes_preview}…" if len(entry["notes"]) > 80
                  else f"    {'notes':<14} {notes_preview}")

    print()
    ans = input(f"  Convert all {len(rows):,} rows? [Y / n]: ").strip().lower()
    if ans in ("n", "no"):
        print("  Cancelled. No files were changed.\n")
        return

    # ── Convert ──────────────────────────────────────────────────────────
    entries = []
    skipped = 0
    for seq, row in enumerate(rows):
        entry = row_to_entry(row, mapping, seq)
        if not entry["company"] and not entry["position"]:
            skipped += 1
            continue
        entries.append(entry)

    if skipped:
        print(f"\n  Skipped {skipped} rows with no company or position.")

    # ── Merge or replace ─────────────────────────────────────────────────
    existing: list[dict] = []
    if Path(APPLICATIONS_FILE).exists() and not replace:
        try:
            with open(APPLICATIONS_FILE, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            existing = []

        print(f"\n  applications.json already has {len(existing):,} entries.")
        print(f"  Imported entries: {len(entries):,}")
        print()
        print("  Options:")
        print("    1. Merge  — add imported rows to existing data (default)")
        print("    2. Replace — overwrite existing data entirely")
        print("    3. Cancel")
        choice = input("\n  Choice [1]: ").strip() or "1"
        if choice == "3":
            print("  Cancelled. No files were changed.\n")
            return
        if choice == "2":
            existing = []

    # Imported entries go at the front (most-recently added on top)
    combined = entries + existing

    with open(APPLICATIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(combined, f, indent=2, ensure_ascii=False)

    # ── Summary ──────────────────────────────────────────────────────────
    print("\n" + "═" * 64)
    print(f"  ✓  Written {len(combined):,} total entries to {APPLICATIONS_FILE}")
    if existing:
        print(f"     {len(entries):,} imported  +  {len(existing):,} existing")
    print()

    # Status breakdown
    counts = Counter(e["status"] for e in entries)
    print("  Status breakdown of imported rows:")
    for status in STATUSES:
        if status in counts:
            bar = "█" * counts[status]
            print(f"    {status:<16} {counts[status]:>4}  {bar}")

    # Warn about rows whose status could not be mapped
    unmapped_statuses = Counter(
        normalize_cell(row[mapping["status"]]) if "status" in mapping and mapping["status"] < len(row) else ""
        for row in rows
    )
    unknown = {k: v for k, v in unmapped_statuses.items()
               if k and normalize_status(k) == "Applied" and k.lower() != "applied"}
    if unknown:
        print()
        print("  Note: the following status values were mapped to 'Applied'")
        print("  (they didn't match any known pattern — edit manually if needed):")
        for val, count in sorted(unknown.items(), key=lambda x: -x[1]):
            print(f"    {val!r:<30}  {count}×")

    print()


if __name__ == "__main__":
    main()
