#!/usr/bin/env python3
"""
Job Application Assistant — GUI v3
Redesigned interface: dark navy, coral accent, clean Anthropic-inspired typography.
"""
from __future__ import annotations

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading, csv, re
from pathlib import Path
import os, json
from datetime import datetime

from dotenv import load_dotenv
from job_application_assistant import JobApplicationAssistant
from import_tracker import read_csv, detect_mapping, row_to_entry

load_dotenv()

# ─────────────────────────────────────────────────────────────────────────────
# Design tokens  (Anthropic-inspired dark navy + coral accent)
# ─────────────────────────────────────────────────────────────────────────────
C = {
    # Backgrounds — layered depth
    "bg":           "#0f1117",   # near-black navy  — main canvas
    "surface":      "#181b26",   # dark navy        — cards, panels
    "surface2":     "#1e2233",   # lighter navy     — hover, secondary
    "surface3":     "#252840",   # lighter still    — active / inputs inside cards
    "input_bg":     "#0b0d16",   # darkest          — text areas and entries
    "sidebar_bg":   "#13151f",   # sidebar (slightly distinct from bg)

    # Accent — warm coral (Anthropic signature)
    "accent":       "#e07640",   # coral orange     — primary CTAs, active nav
    "accent_dim":   "#b85c2e",   # darker           — hover / pressed
    "accent_soft":  "#f0a070",   # lighter          — highlights, selection fg

    # Secondary accent — deep forest green
    "green":        "#2e7d5a",   # forest green     — success actions
    "green_dim":    "#1a4c36",   # dark green bg

    # Semantic
    "success":      "#3aaa6e",   # readable green
    "success_dim":  "#0d2018",   # green badge bg
    "warning":      "#c8961e",   # amber
    "danger":       "#d44e4e",   # red
    "danger_dim":   "#1c0808",   # dark red bg

    # Text
    "text":         "#eceef5",   # near-white, slightly cool — primary
    "text_dim":     "#868fa8",   # medium gray               — secondary
    "text_muted":   "#4e5268",   # dim                       — labels/hints
    "text_inv":     "#0f1117",   # dark text on light bg

    # Structure
    "border":       "#20243a",   # very subtle card borders
    "border_bright":"#32374f",   # visible borders / focus rings
    "divider":      "#181b28",   # section dividers
}

C_DARK = C.copy()   # keep a reference to the original dark palette

C_LIGHT = {
    "bg":           "#f4f5f7",
    "surface":      "#ffffff",
    "surface2":     "#eaecf0",
    "surface3":     "#dde0e8",
    "input_bg":     "#ffffff",
    "sidebar_bg":   "#f0f1f4",
    "accent":       "#c0582a",
    "accent_dim":   "#9e4520",
    "accent_soft":  "#d4703c",
    "green":        "#2e7d5a",
    "green_dim":    "#d4eee2",
    "success":      "#1e7d4a",
    "success_dim":  "#d4f0e4",
    "warning":      "#a07010",
    "danger":       "#c0303030",
    "danger_dim":   "#fde8e8",
    "text":         "#1a1c26",
    "text_dim":     "#505468",
    "text_muted":   "#8890a8",
    "text_inv":     "#ffffff",
    "border":       "#d8dae4",
    "border_bright":"#b8bcd0",
    "divider":      "#e4e6ec",
}

FF = "Segoe UI"
FM = "Consolas"
HISTORY_FILE          = "history.json"
APPLICATIONS_FILE     = "applications.json"
CONFIG_FILE           = "config.json"
PROFILE_FILE          = "profile.md"
PROFILE_INSTRUCTIONS_FILE = "profile_instructions.md"
PROFILE_PERSONAL_FILE = "profile_personal.md"

NAV_ITEMS = [
    ("⚙", "Setup"),
    ("◎", "Evaluate"),
    ("✦", "Generate"),
    ("◇", "Q & A"),
    ("▣", "Package"),
    ("◉", "Interview"),
    ("☰", "History"),
    ("◐", "Profile"),
    ("▤", "Tracker"),
    ("◑", "Stats"),
]

STATUS_COLORS = {
    "Applied":   "#82c4e8",
    "Interview": "#e07640",
    "Offer":     "#3aaa6e",
    "Rejected":  "#d44e4e",
    "":          "#4e5268",
}

MODEL_COSTS = {
    "gpt-4o":        0.005,
    "gpt-4-turbo":   0.016,
    "gpt-4":         0.040,
    "gpt-3.5-turbo": 0.001,
    "gpt-5.2":       0.020,
}

STATUSES = [
    "Watching", "Applied", "Phone Screen", "Interview",
    "Final Round", "Offer", "Accepted", "Rejected", "Withdrawn",
]

# Ordinal rank for forward-progression tracking.
# Terminal outcomes (Rejected/Withdrawn) don't advance peak_stage.
STAGE_ORDER = {
    "Watching":     0,
    "Applied":      1,
    "Phone Screen": 2,
    "Interview":    3,
    "Final Round":  4,
    "Offer":        5,
    "Accepted":     6,
}
# Statuses that represent a terminal outcome (don't advance the funnel)
TERMINAL_STATUSES = {"Rejected", "Withdrawn"}

# Very subtle row tints — visual identity comes from coloured "● Status" text
STATUS_ROW_COLORS = {
    "Watching":     ("#131520", "#565870"),
    "Applied":      ("#111928", "#6ab0d8"),
    "Phone Screen": ("#0f1a2e", "#8ac8ee"),
    "Interview":    ("#1c1508", "#c8921e"),
    "Final Round":  ("#1e1206", "#d4920c"),
    "Offer":        ("#0a1c10", "#3aaa6e"),
    "Accepted":     ("#071a0e", "#3ec870"),
    "Rejected":     ("#1c0808", "#cc5050"),
    "Withdrawn":    ("#121420", "#404458"),
}

TRACKER_COLS = [
    ("date",       "Date Applied",  95),
    ("company",    "Company",      135),
    ("position",   "Position",     155),
    ("location",   "Location",     105),
    ("status",     "Status",       110),
    ("salary_jd",  "Salary (JD)",  110),
    ("salary_req", "Requested",     92),
    ("source",     "Source",        95),
]


# ─────────────────────────────────────────────────────────────────────────────
# Profile Setup Wizard
# ─────────────────────────────────────────────────────────────────────────────

class ProfileWizard(tk.Toplevel):
    """Guided multi-step form for generating profile_personal.md from scratch."""

    STEPS = [
        "Identity & Background",
        "Education",
        "Technical Skills",
        "Career Goals & Positioning",
        "Experience & Key Projects",
        "Salary & Awards",
        "Career Narrative & Why Now",
        "Domains of Interest & Target Companies",
        "Honest Gaps & STAR Stories",
        "Search Logistics & Portfolio",
    ]

    def __init__(self, parent, on_save=None):
        super().__init__(parent)
        self.on_save  = on_save
        self._step    = 0
        self._frames  = []   # one Frame per step
        self._widgets = {}   # keyed by field name → widget

        self.title("Profile Setup Wizard")
        self.configure(bg=C["bg"])
        self.geometry("680x620")
        self.resizable(False, True)
        self.grab_set()

        # ── Header ────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C["surface"], pady=14)
        hdr.pack(fill="x")
        self._step_lbl = tk.Label(hdr, text="", font=(FF, 10, "bold"),
                                  bg=C["surface"], fg=C["text"])
        self._step_lbl.pack(side="left", padx=20)
        self._progress_lbl = tk.Label(hdr, text="", font=(FF, 8),
                                      bg=C["surface"], fg=C["text_muted"])
        self._progress_lbl.pack(side="right", padx=20)

        # ── Scrollable content area ────────────────────────────────────────
        self._canvas = tk.Canvas(self, bg=C["bg"], highlightthickness=0)
        self._canvas.pack(fill="both", expand=True)
        sb = tk.Scrollbar(self._canvas, orient="vertical",
                          command=self._canvas.yview,
                          bg=C["surface2"], troughcolor=C["sidebar_bg"], relief="flat")
        sb.pack(side="right", fill="y")
        self._canvas.configure(yscrollcommand=sb.set)
        self._canvas.bind("<Configure>",
                          lambda e: self._canvas.configure(
                              scrollregion=self._canvas.bbox("all")))
        self._canvas.bind_all("<MouseWheel>",
                              lambda e: self._canvas.yview_scroll(
                                  int(-1*(e.delta/120)), "units"))

        self._content = tk.Frame(self._canvas, bg=C["bg"])
        self._content_win = self._canvas.create_window(
            (0, 0), window=self._content, anchor="nw")
        self._canvas.bind("<Configure>",
                          lambda e: self._canvas.itemconfig(
                              self._content_win, width=e.width))

        # ── Navigation bar ─────────────────────────────────────────────────
        nav = tk.Frame(self, bg=C["surface"], pady=12)
        nav.pack(fill="x", side="bottom")
        self._back_btn = tk.Button(nav, text="← Back", font=(FF, 9),
                                   bg=C["surface2"], fg=C["text"],
                                   relief="flat", bd=0, padx=16, pady=6,
                                   cursor="hand2", command=self._back,
                                   activebackground=C["surface3"],
                                   activeforeground=C["text"])
        self._back_btn.pack(side="left", padx=(16, 8))
        self._next_btn = tk.Button(nav, text="Next →", font=(FF, 9, "bold"),
                                   bg=C["accent"], fg=C["text_inv"],
                                   relief="flat", bd=0, padx=20, pady=6,
                                   cursor="hand2", command=self._next,
                                   activebackground=C["accent_dim"],
                                   activeforeground=C["text_inv"])
        self._next_btn.pack(side="right", padx=(8, 16))
        tk.Label(nav, text="All fields are optional — fill what you know.",
                 font=(FF, 7), bg=C["surface"], fg=C["text_muted"]).pack(side="right", padx=8)

        self._build_all_steps()
        self._show_step(0)

    # ── Widget helpers ─────────────────────────────────────────────────────

    def _field(self, parent, label, key, height=1, hint=""):
        """Create a labelled entry (height=1) or text area (height>1)."""
        row = tk.Frame(parent, bg=C["bg"])
        row.pack(fill="x", pady=(0, 8))
        lbl = tk.Label(row, text=label, font=(FF, 8, "bold"),
                       bg=C["bg"], fg=C["text_dim"], anchor="w")
        lbl.pack(anchor="w")
        if hint:
            tk.Label(row, text=hint, font=(FF, 7), bg=C["bg"],
                     fg=C["text_muted"], anchor="w").pack(anchor="w")
        border = tk.Frame(row, bg=C["border_bright"], padx=1, pady=1)
        border.pack(fill="x")
        if height == 1:
            var = tk.StringVar()
            w = tk.Entry(border, textvariable=var, font=(FF, 9),
                         bg=C["input_bg"], fg=C["text"],
                         insertbackground=C["accent_soft"],
                         relief="flat", bd=0, highlightthickness=0)
            w.pack(fill="x", ipady=5, padx=6)
            w.bind("<FocusIn>",  lambda e: border.configure(bg=C["accent"]))
            w.bind("<FocusOut>", lambda e: border.configure(bg=C["border_bright"]))
            self._widgets[key] = var
        else:
            w = tk.Text(border, height=height, font=(FF, 9),
                        bg=C["input_bg"], fg=C["text"],
                        insertbackground=C["accent_soft"],
                        relief="flat", bd=0, wrap="word",
                        padx=6, pady=4)
            w.pack(fill="x")
            w.bind("<FocusIn>",  lambda e: border.configure(bg=C["accent"]))
            w.bind("<FocusOut>", lambda e: border.configure(bg=C["border_bright"]))
            self._widgets[key] = w
        return w

    def _section_lbl(self, parent, text):
        tk.Label(parent, text=text, font=(FF, 9, "bold"),
                 bg=C["bg"], fg=C["accent"]).pack(anchor="w", pady=(12, 4))

    def _get(self, key) -> str:
        w = self._widgets.get(key)
        if w is None:
            return ""
        if isinstance(w, tk.StringVar):
            return w.get().strip()
        return w.get(1.0, "end").strip()

    def _bullets(self, key) -> str:
        """Return multi-line text as bullet list, each non-empty line prefixed with '- '."""
        lines = [ln.strip() for ln in self._get(key).splitlines() if ln.strip()]
        return "\n".join(f"- {ln}" for ln in lines) if lines else "- [not provided]"

    # ── Step builders ──────────────────────────────────────────────────────

    def _build_all_steps(self):
        for i in range(len(self.STEPS)):
            f = tk.Frame(self._content, bg=C["bg"], padx=24, pady=20)
            self._frames.append(f)
            builder = [self._step0, self._step1, self._step2,
                       self._step3, self._step4, self._step5,
                       self._step6, self._step7, self._step8,
                       self._step9][i]
            builder(f)

    def _step0(self, f):
        self._section_lbl(f, "Who are you?")
        self._field(f, "Full name", "name", hint="e.g.  Jane Smith")
        self._field(f, "Current role & employer", "current_role",
                    hint="e.g.  Senior ML Engineer at Acme Corp")
        self._field(f, "Location", "location", hint="e.g.  London, UK")
        self._field(f, "Work authorization", "work_auth",
                    hint='e.g.  "UK Citizen" or "US H-1B — sponsorship required"')
        self._field(f, "Career stage summary", "career_stage",
                    hint='e.g.  "Senior researcher transitioning into industry ML roles"')

    def _step1(self, f):
        self._section_lbl(f, "Education")
        self._field(f, "Highest degree", "edu1",
                    hint="Degree · Field · University · Year  (+ any fellowships or distinctions)")
        self._field(f, "Second degree (optional)", "edu2")
        self._field(f, "Third degree (optional)", "edu3")
        self._field(f, "Academic highlights / publication metrics (optional)", "edu_notes",
                    height=3, hint="e.g. Top 5% of cohort · 200+ citations · exchange at MIT")

    def _step2(self, f):
        self._section_lbl(f, "Technical Skills")
        self._field(f, "Primary strengths — one per line", "strengths", height=5,
                    hint="List 5–8 specific, verifiable skills")
        self._field(f, "Domain expertise — one per line", "domains", height=3,
                    hint="e.g.  Earth Observation  /  NLP  /  Computer Vision")
        self._field(f, "Tools & ecosystems — one per line", "tools", height=3,
                    hint="e.g.  Python · PyTorch · AWS  /  Kubernetes · MLflow")
        self._field(f, "Languages & proficiency", "languages",
                    hint="e.g.  English (native)  ·  Spanish (B2)  ·  French (A2)")

    def _step3(self, f):
        self._section_lbl(f, "Career Goals & Positioning")
        self._field(f, "Target role types — one per line", "target_roles", height=3,
                    hint="e.g.  Applied ML Engineer  /  Research Scientist  /  AI Lead")
        self._field(f, "Target company types — one per line", "target_companies", height=2,
                    hint="e.g.  Deep tech startups  /  Research labs  /  Scale-ups")
        self._field(f, "Location preferences", "loc_prefs",
                    hint="e.g.  London preferred · open to remote in EU")
        self._field(f, "What you value in a role", "role_values", height=2,
                    hint="e.g.  Depth over buzzwords · ownership · research-to-production")
        self._field(f, "Key priorities (financial, impact, lifestyle…)", "priorities", height=2)
        self._field(f, "Narrative to manage — the gap to bridge", "narrative", height=3,
                    hint='e.g.  "Academic background must read as production-ready, not over-theoretical"')

    def _step4(self, f):
        self._section_lbl(f, "Experience & Key Projects")
        self._field(f, "Current role — full description", "role_current", height=2,
                    hint="e.g.  Research Scientist, DeepMind (2022–present)")
        self._field(f, "Previous roles — one per line", "roles_prev", height=3)
        self._section_lbl(f, "Key Projects (up to 3)")
        for i in range(1, 4):
            self._field(f, f"Project {i} — name", f"proj{i}_name")
            self._field(f, f"Project {i} — description", f"proj{i}_desc", height=2,
                        hint="What was built · what impact it had · what's transferable")
        self._field(f, "Publications / output record", "publications", height=2,
                    hint="e.g.  8 papers · 400 citations · 2 NeurIPS · 1 Nature")

    def _step5(self, f):
        self._section_lbl(f, "Salary Benchmarks")
        self._field(f, "Market / location context", "salary_market",
                    hint='e.g.  "London, UK — 2026 market"')
        self._field(f, "Primary target role", "salary_role_a",
                    hint="e.g.  Applied ML Engineer")
        self._field(f, "Salary range for primary role", "salary_range_a",
                    hint="e.g.  £90k–£130k base + equity  (startups: lower cash, more equity)")
        self._field(f, "Secondary role (optional)", "salary_role_b")
        self._field(f, "Salary range for secondary role", "salary_range_b")
        self._field(f, "Minimum acceptable", "salary_min",
                    hint="e.g.  £75k — below this, decline")
        self._field(f, "Fair target range", "salary_fair",
                    hint="e.g.  £90k–£110k")
        self._field(f, "Stretch target", "salary_stretch",
                    hint="e.g.  £120k+")
        self._section_lbl(f, "Fellowships, Awards & Certificates")
        self._field(f, "List — one per line", "awards", height=4,
                    hint="e.g.  Google PhD Fellowship · ICLR Outstanding Paper 2023")

    def _step6(self, f):
        self._section_lbl(f, "Career Narrative & Why Now")
        self._field(f, "Your career narrative — the connecting thread", "narrative_thread", height=5,
                    hint='What links your roles/decisions into one story? e.g. "Started in X, moved to Y because [reason], '
                         'now heading toward Z. From measuring the world to modelling it."')
        self._field(f, "Why industry / why now?", "why_industry", height=4,
                    hint='Your genuine reason for leaving academia / current role. Be specific.\n'
                         'e.g. "I want a faster feedback loop between building and real-world impact."')
        self._field(f, "One-line positioning statement (optional)", "positioning_line",
                    hint='e.g. "Geospatial ML researcher who builds systems that work on real, noisy satellite data"')

    def _step7(self, f):
        self._section_lbl(f, "Domains of Genuine Interest")
        tk.Label(f, text="Only include domains where interest is genuine — the AI uses these for cover letter hooks.",
                 font=(FF, 7), bg=C["bg"], fg=C["text_muted"], wraplength=580).pack(anchor="w", pady=(0, 8))
        self._field(f, "Primary domain (strongest fit)", "domain1", height=2,
                    hint='e.g. "Climate and environmental intelligence — my current work is directly here"')
        self._field(f, "Secondary domain", "domain2", height=2,
                    hint='e.g. "Autonomous systems / robotics — transferable: multimodal fusion, spatial reasoning"')
        self._field(f, "Tertiary domain (optional)", "domain3", height=2)
        self._field(f, "Fourth domain (optional, note if limited experience)", "domain4", height=2)
        self._field(f, "General principle — what kind of problems attract you?", "domain_principle",
                    hint='e.g. "Domains where ML interacts with complex, noisy real-world data and high-stakes decisions"')
        self._field(f, "Specific companies currently on radar", "target_companies_specific", height=3,
                    hint='e.g. "DeepMind, Wayve, Planet Labs, Palantir — companies building real ML systems at scale"')

    def _step8(self, f):
        self._section_lbl(f, "Honest Gaps")
        tk.Label(f, text="Be specific — this prevents the AI from overclaiming in materials.",
                 font=(FF, 7), bg=C["bg"], fg=C["text_muted"], wraplength=580).pack(anchor="w", pady=(0, 8))
        self._field(f, "Primary technical gap", "gap1", height=3,
                    hint='Name it + bridge it. e.g. "No Docker/Kubernetes/CI-CD. Bridge: I own the full ML pipeline; '
                         'the infrastructure layer is where I\'m growing."')
        self._field(f, "Secondary gap", "gap2", height=3,
                    hint='e.g. "Limited direct industry experience. Bridge: built production-grade systems in academic settings '
                         'with external stakeholders (DLR, Airbus)."')
        self._field(f, "Hard constraint to always flag", "hard_constraint",
                    hint='e.g. "Cannot obtain UK defence clearance — decline any role requiring it"')
        self._section_lbl(f, "STAR Stories (behavioral interview answers)")
        self._field(f, "Story 1 — Title & Summary", "star1", height=4,
                    hint="Situation · Task · Action · Result. Give real names and numbers.\n"
                         'e.g. "Convincing non-ML ecologists to adopt deep learning. Designed interpretable '
                         'architecture → 80% bias reduction → team converted."')
        self._field(f, "Story 2 — Title & Summary", "star2", height=4,
                    hint="e.g. Multi-team technical delivery / project under constraints")
        self._field(f, "Story 3 — Title & Summary", "star3", height=4,
                    hint="e.g. Handling noisy/messy data, debugging a hard failure, influencing without authority")

    def _step9(self, f):
        self._section_lbl(f, "Search Logistics")
        self._field(f, "Contract end / availability", "availability",
                    hint='e.g. "Contract ends March 2027; 3-month notice, flexible with manager"')
        self._field(f, "Notice period", "notice",
                    hint='e.g. "3 months (negotiable downward for the right role)"')
        self._field(f, "Work mode preference", "work_mode",
                    hint='e.g. "Flexible — office, hybrid, or remote all acceptable"')
        self._field(f, "Hard location constraint", "location_constraint",
                    hint='e.g. "Must remain in UK (visa-tied)"')
        self._field(f, "Sectors / roles to avoid (and why)", "anti_goals", height=3,
                    hint='e.g. "Defence (clearance unavailable). Pure data engineering / dashboard roles with no modelling."')
        self._field(f, "Current salary", "current_salary",
                    hint='e.g. "£42,000/year — any offer should represent a meaningful step up"')
        self._section_lbl(f, "Public Portfolio & GitHub")
        self._field(f, "Portfolio / personal website URL", "portfolio_url",
                    hint="e.g. https://yourname.github.io/")
        self._field(f, "Key GitHub repos — one per line", "github_repos", height=4,
                    hint="Name: What it is and why it's worth citing\n"
                         'e.g. "cvnnTorch: PyTorch package for complex-valued networks — shows ability to package research as tooling"')

    # ── Navigation ─────────────────────────────────────────────────────────

    def _show_step(self, idx):
        for f in self._frames:
            f.pack_forget()
        self._frames[idx].pack(fill="both", expand=True)
        self._canvas.yview_moveto(0)

        self._step_lbl.configure(text=f"Step {idx+1}: {self.STEPS[idx]}")
        self._progress_lbl.configure(text=f"{idx+1} / {len(self.STEPS)}")
        self._back_btn.configure(state="normal" if idx > 0 else "disabled")
        is_last = idx == len(self.STEPS) - 1
        self._next_btn.configure(text="✓  Save Profile" if is_last else "Next →")

    def _next(self):
        if self._step < len(self.STEPS) - 1:
            self._step += 1
            self._show_step(self._step)
        else:
            self._save()

    def _back(self):
        if self._step > 0:
            self._step -= 1
            self._show_step(self._step)

    # ── Markdown generation ────────────────────────────────────────────────

    def _save(self):
        md = self._generate_markdown()
        path = Path(__file__).parent / PROFILE_PERSONAL_FILE
        try:
            path.write_text(md, encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save profile:\n{e}", parent=self)
            return
        messagebox.showinfo("Saved",
                            f"Personal profile saved to {PROFILE_PERSONAL_FILE}.\n\n"
                            "Go to the Profile page to review and refine it.",
                            parent=self)
        if self.on_save:
            self.on_save()
        self.destroy()

    def _generate_markdown(self) -> str:
        g = self._get

        def edu_lines():
            lines = [f"- {g(k)}" for k in ("edu1", "edu2", "edu3") if g(k)]
            return "\n".join(lines) if lines else "- [not provided]"

        def proj_block():
            blocks = []
            for i in range(1, 4):
                name = g(f"proj{i}_name")
                desc = g(f"proj{i}_desc")
                if name or desc:
                    blocks.append(f"- {name or '[Project name]'}: {desc or '[Description]'}")
            return "\n".join(blocks) if blocks else "- [not provided]"

        def award_lines():
            lines = [ln.strip() for ln in g("awards").splitlines() if ln.strip()]
            return "\n".join(f"- {ln}" for ln in lines) if lines else "- [not provided]"

        def domain_block():
            parts = []
            labels = ["1.", "2.", "3.", "4."]
            keys   = ["domain1", "domain2", "domain3", "domain4"]
            for lbl, key in zip(labels, keys):
                v = g(key)
                if v:
                    parts.append(f"{lbl} {v}")
            return "\n\n".join(parts) if parts else "[not provided]"

        def star_block():
            parts = []
            for i, key in enumerate(["star1", "star2", "star3"], 1):
                v = g(key)
                if v:
                    parts.append(f"STORY {i}: {v}")
            return "\n\n".join(parts) if parts else "[not provided]"

        def github_block():
            lines = [ln.strip() for ln in g("github_repos").splitlines() if ln.strip()]
            return "\n".join(f"- {ln}" for ln in lines) if lines else "- [not provided]"

        salary_a = g("salary_role_a") or "[Role Type A]"
        salary_b = g("salary_role_b") or "[Role Type B]"

        return f"""1. IDENTITY AND BACKGROUND

Name: {g("name") or "[YOUR FULL NAME]"}
Current role: {g("current_role") or "[YOUR CURRENT JOB TITLE AND EMPLOYER]"}
Location: {g("location") or "[YOUR CITY, COUNTRY]"}
Work authorization: {g("work_auth") or "[YOUR WORK AUTHORIZATION STATUS]"}
Career stage: {g("career_stage") or "[YOUR CAREER STAGE DESCRIPTION]"}
Website / Portfolio: {g("portfolio_url") or "[YOUR PORTFOLIO URL]"}

Education:
{edu_lines()}

{g("edu_notes") or ""}


2. TECHNICAL SKILL PROFILE

Primary strengths (core differentiators — weight these heavily):
{self._bullets("strengths")}

Domain expertise:
{self._bullets("domains")}

Tools and ecosystems:
{self._bullets("tools")}

Languages: {g("languages") or "[YOUR LANGUAGES AND PROFICIENCY LEVELS]"}


3. CAREER INTENT AND DIRECTION

Short- to mid-term goal:
{self._bullets("target_roles")}
{self._bullets("target_companies")}

Preferences:
- {g("loc_prefs") or "[LOCATION PREFERENCES]"}
- {g("work_mode") or "[WORK MODE PREFERENCE]"}
- {g("role_values") or "[ROLE CHARACTERISTICS YOU VALUE]"}
- {g("priorities") or "[OTHER PRIORITIES]"}

Contract / availability: {g("availability") or "[WHEN YOU CAN START / NOTICE PERIOD]"}

Anti-goals:
- {g("anti_goals") or "[SECTORS OR ROLE TYPES TO AVOID]"}

Identity tension to manage carefully:
- {g("narrative") or "[DESCRIBE THE POSITIONING CHALLENGE — e.g. academic background must read as production-ready]"}


4. TARGET ROLE TYPES AND POSITIONING STRATEGIES

Based on the target roles above, decide which archetype fits each role and adapt accordingly.
Refer to Section 3 for your stated targets and use the profile to construct the best angle.

(Refine this section manually to add role-specific positioning strategies as needed.)


5. SALARY BENCHMARKS — {g("salary_market") or "[YOUR MARKET/LOCATION]"} (UPDATE ANNUALLY)

Current salary: {g("current_salary") or "[YOUR CURRENT SALARY — use as negotiation baseline]"}

Use these as anchoring data when estimating expected salary. Always give a specific range with reasoning.

{salary_a}:
- {g("salary_range_a") or "[SALARY RANGE]"}

{salary_b}:
- {g("salary_range_b") or "[SALARY RANGE]"}

Realistic anchor range:
- Minimum acceptable:  {g("salary_min") or "[YOUR MINIMUM]"}
- Fair target:         {g("salary_fair") or "[YOUR TARGET RANGE]"}
- Stretch target:      {g("salary_stretch") or "[YOUR STRETCH RANGE]"}

When estimating salary:
- Always give a specific range with reasoning — never vague
- Explain: company stage, location premium, role specificity, seniority signals
- Flag clearly if the role appears underpaying for the candidate's level


6. FELLOWSHIPS, AWARDS, AND CERTIFICATES

{award_lines()}


7. CV KNOWLEDGE (authoritative source — use for specific project details)

Research/work roles (most recent first):
- {g("role_current") or "[CURRENT ROLE, ORGANISATION, DATES]"}
{self._bullets("roles_prev")}

Key projects (reference these specifically):
{proj_block()}

Publication/output record: {g("publications") or "[CITATION COUNT, VENUES, DATASETS, NOTABLE IMPACT]"}


8. CAREER NARRATIVE

The core thread (use this framing for "tell me about yourself" and cover letter openers):
{g("narrative_thread") or "[YOUR CAREER NARRATIVE — the connecting thread across your roles and decisions]"}

{("One-line positioning: " + g("positioning_line")) if g("positioning_line") else ""}

Why industry / why now (use when asked "why leave current role?"):
{g("why_industry") or "[YOUR GENUINE REASON — specific and honest, not 'I want new challenges']"}

Do NOT use: "I want new challenges" / "I'm passionate about industry" — these are weak. Use the specific framing above.


9. DOMAINS OF GENUINE INTEREST

Only include domains where interest is genuine — do not fabricate enthusiasm.

{domain_block()}

General principle: {g("domain_principle") or "[WHAT KIND OF PROBLEMS OR ENVIRONMENTS ATTRACT YOU]"}

Companies currently on radar: {g("target_companies_specific") or "[SPECIFIC COMPANIES YOU ARE INTERESTED IN]"}

Cover letter guidance: Never use generic "I'm excited to bring my skills to [Company]" — always connect to something specific about what they build or the problem they're solving.


10. HONEST GAPS AND HOW TO ADDRESS THEM

Gap 1: {g("gap1") or "[YOUR PRIMARY GAP — name it and state the bridge response]"}

Gap 2: {g("gap2") or "[YOUR SECONDARY GAP AND BRIDGE]"}

Hard constraint: {g("hard_constraint") or "[ANY NON-NEGOTIABLE CONSTRAINT — e.g. cannot obtain clearance for X]"}


11. BEST STAR STORIES

{star_block()}


12. SEARCH LOGISTICS AND CONSTRAINTS

Contract/availability: {g("availability") or "[WHEN YOU CAN START]"}
Notice period: {g("notice") or "[NOTICE PERIOD]"}
Work mode: {g("work_mode") or "[WORK MODE PREFERENCE]"}
Location constraint: {g("location_constraint") or "[HARD LOCATION CONSTRAINT IF ANY]"}
Current salary: {g("current_salary") or "[YOUR CURRENT SALARY]"}

Sectors/roles to avoid:
{g("anti_goals") or "[SECTORS AND ROLE TYPES TO AVOID — AND WHY]"}


13. PUBLIC PORTFOLIO AND GITHUB

Website: {g("portfolio_url") or "[YOUR PORTFOLIO URL]"}

Notable public repositories:
{github_block()}
"""


# ─────────────────────────────────────────────────────────────────────────────
# Application
# ─────────────────────────────────────────────────────────────────────────────
class JobAssistantV3:

    def __init__(self, root):
        self.root = root
        self.root.title("Job Application Assistant")
        self.root.geometry("1440x880")
        self.root.minsize(1100, 700)
        self.root.configure(bg=C["bg"])

        # ── State ─────────────────────────────────────────────────────────
        self.assistant        = None
        self.cv_profiles      = []    # [{"name": str, "path": str}, ...]
        self.active_cv_name   = ""
        self.cv_path          = tk.StringVar(value=os.getenv("CV_PATH", "cv.pdf"))
        # Ollama / provider state
        self.provider         = "openai"   # "openai" | "ollama"
        self.ollama_url       = "http://localhost:11434"
        self.ollama_model     = ""
        self._ollama_url_var  = tk.StringVar(value="http://localhost:11434")
        self.model_var        = tk.StringVar(value="gpt-4o")
        self._font_size       = 10    # default; _load_config may override
        self._load_config()           # may override cv_path, model, font_size from saved config
        self.company_name       = tk.StringVar()
        self.role_title         = tk.StringVar()
        self.recruiter_name     = tk.StringVar()
        self._followup_name_var = tk.StringVar()
        self._session_tokens = 0
        self._session_cost   = 0.0
        self._pkg_running    = False
        self._shared_jd      = ""
        self._current_page   = "Setup"
        self.history         = self._load_history()

        self._hist_search_var      = tk.StringVar()
        self._hist_filter_status   = ""
        self._hist_display_indices = []

        self.applications          = self._load_applications()
        self._tracker_selected_id  = None
        self._tracker_sort_col     = None
        self._tracker_sort_rev     = False
        self._tracker_filter_status = None
        self._stats_count_labels   = {}
        self._trk_company          = tk.StringVar()
        self._trk_position         = tk.StringVar()
        self._trk_location         = tk.StringVar()
        self._trk_date             = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self._trk_source           = tk.StringVar()
        self._trk_status_var       = tk.StringVar(value="Applied")
        self._trk_salary_jd        = tk.StringVar()
        self._trk_salary_req       = tk.StringVar()
        self._trk_date_interview   = tk.StringVar()
        self._trk_date_decision    = tk.StringVar()
        self._trk_peak_stage       = tk.StringVar(value="")

        # ── Build UI ──────────────────────────────────────────────────────
        self._build_header()
        self._build_body()
        self._build_statusbar()
        self._setup_keyboard_shortcuts()
        self._nav_to("Setup")
        self.root.after(400, self._auto_init)

    # ═════════════════════════════════════════════════════════════════════════
    # Layout builders
    # ═════════════════════════════════════════════════════════════════════════

    def _build_header(self):
        hdr = tk.Frame(self.root, bg=C["surface"], height=54)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        # Left: brand
        brand = tk.Frame(hdr, bg=C["surface"])
        brand.pack(side="left", fill="y")
        tk.Label(brand, text="◈", font=(FF, 14), bg=C["surface"],
                 fg=C["accent"]).pack(side="left", padx=(20, 6), pady=14)
        tk.Label(brand, text="Job Application Assistant", font=(FF, 12, "bold"),
                 bg=C["surface"], fg=C["text"]).pack(side="left", pady=14)
        # Right: font size controls + init status
        self._init_dot = tk.Label(hdr, text="● Not initialized",
                                  font=(FF, 8), bg=C["surface"], fg=C["danger"])
        self._init_dot.pack(side="right", padx=(0, 20))
        self._theme_btn = tk.Button(hdr, text="☀", font=(FF, 10), bg=C["surface2"],
                                    fg=C["text_muted"], relief="flat", bd=0,
                                    padx=8, pady=4, cursor="hand2",
                                    activebackground=C["surface3"],
                                    activeforeground=C["text"],
                                    command=self._toggle_theme)
        self._theme_btn.pack(side="right", padx=(0, 8))
        for symbol, delta in (("A+", 1), ("A−", -1)):
            tk.Button(hdr, text=symbol, font=(FF, 8), bg=C["surface2"], fg=C["text_muted"],
                      relief="flat", bd=0, padx=8, pady=4, cursor="hand2",
                      activebackground=C["surface3"], activeforeground=C["text"],
                      command=lambda d=delta: self._adjust_font_size(d)).pack(
                side="right", padx=(0, 4))
        # Coral accent line
        tk.Frame(self.root, bg=C["accent"], height=1).pack(fill="x")

    def _build_body(self):
        body = tk.Frame(self.root, bg=C["bg"])
        body.pack(fill="both", expand=True)
        self._build_sidebar(body)
        # 1px divider between sidebar and content
        tk.Frame(body, bg=C["border_bright"], width=1).pack(side="left", fill="y")
        self.content = tk.Frame(body, bg=C["bg"])
        self.content.pack(side="left", fill="both", expand=True)

        self.pages = {
            "Setup":     self._page_setup(self.content),
            "Evaluate":  self._page_evaluate(self.content),
            "Generate":  self._page_generate(self.content),
            "Q & A":     self._page_qa(self.content),
            "Package":   self._page_package(self.content),
            "Interview": self._page_interview(self.content),
            "History":   self._page_history(self.content),
            "Profile":   self._page_profile(self.content),
            "Tracker":   self._page_tracker(self.content),
            "Stats":     self._page_stats(self.content),
        }
        self._jd_widgets = {
            "Evaluate":  self.eval_job_text,
            "Generate":  self.gen_job_text,
            "Q & A":     self.qa_job_text,
            "Package":   self.pkg_job_text,
            "Interview": self.interview_job_text,
        }

    def _build_sidebar(self, parent):
        sb = tk.Frame(parent, bg=C["sidebar_bg"], width=196)
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)

        # App label
        tk.Frame(sb, bg=C["divider"], height=1).pack(fill="x", pady=(0, 0))
        tk.Label(sb, text="NAVIGATION", font=(FF, 7, "bold"),
                 bg=C["sidebar_bg"], fg=C["text_muted"]).pack(anchor="w", padx=18, pady=(16, 8))

        self._nav_btns       = {}
        self._nav_indicators = {}

        for icon, name in NAV_ITEMS:
            row = tk.Frame(sb, bg=C["sidebar_bg"])
            row.pack(fill="x", pady=1)
            # Left accent indicator (3px, color changes on active)
            ind = tk.Frame(row, bg=C["sidebar_bg"], width=3)
            ind.pack(side="left", fill="y")
            self._nav_indicators[name] = ind
            # Nav button
            btn = tk.Button(row, text=f"  {icon}   {name}",
                            font=(FF, 10), anchor="w", relief="flat", bd=0,
                            cursor="hand2", bg=C["sidebar_bg"], fg=C["text_muted"],
                            activebackground=C["surface2"], activeforeground=C["text"],
                            padx=10, pady=9,
                            command=lambda n=name: self._nav_to(n))
            btn.pack(side="left", fill="both", expand=True)
            self._nav_btns[name] = btn

    def _build_statusbar(self):
        bar = tk.Frame(self.root, bg=C["sidebar_bg"], height=26)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        tk.Frame(self.root, bg=C["border_bright"], height=1).pack(fill="x", side="bottom")
        self._status_lbl = tk.Label(bar, text="Ready", font=(FF, 8),
                                    bg=C["sidebar_bg"], fg=C["text_muted"], anchor="w")
        self._status_lbl.pack(side="left", padx=14)
        self._token_lbl = tk.Label(bar, text="Session: 0 tokens  |  $0.000",
                                   font=(FF, 8), bg=C["sidebar_bg"], fg=C["text_muted"], anchor="e")
        self._token_lbl.pack(side="right", padx=14)

    # ═════════════════════════════════════════════════════════════════════════
    # Keyboard shortcuts
    # ═════════════════════════════════════════════════════════════════════════

    def _setup_keyboard_shortcuts(self):
        self.root.bind_all("<Control-Return>", lambda e: self._on_ctrl_enter())
        self.root.bind_all("<Control-s>",      lambda e: self._on_ctrl_s())
        for i, (_, name) in enumerate(NAV_ITEMS, 1):
            if i <= 9:   # Ctrl+10 is not a valid key
                self.root.bind_all(f"<Control-Key-{i}>", lambda e, n=name: self._nav_to(n))

    def _on_ctrl_enter(self):
        actions = {
            "Setup":     self._initialize_assistant,
            "Evaluate":  self._evaluate_job_fit,
            "Generate":  self._generate_cv_summary,
            "Q & A":     self._answer_question,
            "Package":   self._generate_package,
            "Interview": self._generate_interview_prep,
            "Profile":   self._save_profile,
            "Tracker":   self._tracker_save_entry,
        }
        action = actions.get(self._current_page)
        if action:
            action()

    def _on_ctrl_s(self):
        if self._current_page == "Profile":
            self._save_profile()
            return
        if self._current_page == "Tracker":
            self._tracker_save_entry()
            return
        output_map = {
            "Evaluate":  self.eval_results,
            "Generate":  self.gen_output,
            "Q & A":     self.qa_output,
            "Package":   self.pkg_output,
            "Interview": self.interview_output,
            "History":   self.hist_content,
        }
        widget = output_map.get(self._current_page)
        if widget:
            self._save_text(widget)

    # ═════════════════════════════════════════════════════════════════════════
    # Navigation
    # ═════════════════════════════════════════════════════════════════════════

    def _nav_to(self, name):
        if self._current_page in self._jd_widgets:
            text = self._jd_widgets[self._current_page].get(1.0, "end").strip()
            if text:
                self._shared_jd = text
        for p in self.pages.values():
            p.pack_forget()
        # Reset all nav buttons
        for n, b in self._nav_btns.items():
            b.configure(bg=C["sidebar_bg"], fg=C["text_muted"])
            self._nav_indicators[n].configure(bg=C["sidebar_bg"])
        # Activate selected
        self.pages[name].pack(fill="both", expand=True)
        self._nav_btns[name].configure(bg=C["surface2"], fg=C["accent"])
        self._nav_indicators[name].configure(bg=C["accent"])
        self._current_page = name
        if name in self._jd_widgets and self._shared_jd:
            target = self._jd_widgets[name]
            if not target.get(1.0, "end").strip():
                target.delete(1.0, "end")
                target.insert(1.0, self._shared_jd)

    # ═════════════════════════════════════════════════════════════════════════
    # Widget factories
    # ═════════════════════════════════════════════════════════════════════════

    def _card(self, parent, title=None, padx=16, pady=12):
        outer = tk.Frame(parent, bg=C["border"], padx=1, pady=1)
        inner = tk.Frame(outer, bg=C["surface"], padx=padx, pady=pady)
        inner.pack(fill="both", expand=True)
        if title:
            tk.Label(inner, text=title, font=(FF, 7, "bold"),
                     bg=C["surface"], fg=C["text_muted"]).pack(anchor="w", pady=(0, 10))
        return outer, inner

    def _btn(self, parent, text, cmd, style="primary", **kw):
        palettes = {
            "primary":   (C["accent"],      C["text_inv"],   C["accent_dim"]),
            "secondary": (C["surface2"],    C["text_dim"],   C["surface3"]),
            "success":   (C["success_dim"], C["success"],    "#071510"),
            "danger":    (C["danger_dim"],  C["danger"],     "#140505"),
            "ghost":     (C["bg"],          C["text_muted"], C["surface2"]),
            "warning":   ("#1e1400",        C["warning"],    "#2c1e00"),
            "outline":   (C["surface"],     C["accent"],     C["surface2"]),
        }
        bg, fg, abg = palettes.get(style, palettes["primary"])
        return tk.Button(parent, text=text, command=cmd, relief="flat", bd=0,
                         cursor="hand2", font=(FF, 9, "bold"),
                         bg=bg, fg=fg, activebackground=abg, activeforeground=fg,
                         padx=14, pady=7, **kw)

    def _textarea(self, parent, height=8, mono=False):
        ff = FM if mono else FF
        t = tk.Text(parent, height=height, font=(ff, 10),
                    bg=C["input_bg"], fg=C["text"],
                    insertbackground=C["accent_soft"],
                    selectbackground=C["surface3"], selectforeground=C["text"],
                    relief="flat", bd=0, wrap="word", padx=12, pady=10, spacing1=2,
                    undo=True, maxundo=60)
        sb = tk.Scrollbar(parent, command=t.yview,
                          bg=C["surface2"], troughcolor=C["sidebar_bg"], relief="flat")
        t.configure(yscrollcommand=sb.set)
        return t, sb

    def _entry(self, parent, textvariable=None, width=30):
        return tk.Entry(parent, textvariable=textvariable, font=(FF, 10),
                        bg=C["input_bg"], fg=C["text"],
                        insertbackground=C["accent_soft"],
                        selectbackground=C["surface3"], selectforeground=C["text"],
                        relief="flat", bd=0,
                        highlightthickness=1,
                        highlightcolor=C["accent"],
                        highlightbackground=C["border"],
                        width=width)

    def _section_header(self, parent, title, subtitle=None):
        """Left-accent-bar section header with clean typography."""
        f = tk.Frame(parent, bg=C["bg"])
        title_row = tk.Frame(f, bg=C["bg"])
        title_row.pack(anchor="w")
        tk.Frame(title_row, bg=C["accent"], width=3).pack(side="left", fill="y", padx=(0, 12))
        tk.Label(title_row, text=title, font=(FF, 17, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(side="left", anchor="s", pady=(0, 2))
        if subtitle:
            tk.Label(f, text=subtitle, font=(FF, 9),
                     bg=C["bg"], fg=C["text_muted"]).pack(anchor="w", pady=(4, 0), padx=(15, 0))
        tk.Frame(f, bg=C["border_bright"], height=1).pack(fill="x", pady=(12, 0))
        return f

    def _lbl(self, parent, text, bg=None, size=9, bold=False, dim=False):
        bg = bg or C["bg"]
        return tk.Label(parent, text=text, font=(FF, size, "bold" if bold else ""),
                        bg=bg, fg=C["text_muted"] if dim else C["text"])

    def _inset_area(self, parent, height=10, mono=False):
        border = tk.Frame(parent, bg=C["border_bright"], padx=1, pady=1)
        inner  = tk.Frame(border, bg=C["input_bg"])
        inner.pack(fill="both", expand=True)
        t, sb = self._textarea(inner, height=height, mono=mono)
        t.pack(side="left", fill="both", expand=True)
        sb.pack(side="left", fill="y")
        t._container = border
        # Focus: highlight border with accent
        t.bind("<FocusIn>",  lambda e: border.configure(bg=C["accent"]))
        t.bind("<FocusOut>", lambda e: border.configure(bg=C["border_bright"]))
        return t, sb

    def _optional_details_card(self, parent, show_recruiter=False):
        opt_outer, opt_card = self._card(parent, "OPTIONAL DETAILS", padx=12, pady=10)
        opt_card.configure(bg=C["surface"])
        fields = [("Company", self.company_name), ("Role Title", self.role_title)]
        if show_recruiter:
            fields.append(("Recruiter", self.recruiter_name))
        for label, var in fields:
            row = tk.Frame(opt_card, bg=C["surface"])
            row.pack(fill="x", pady=2)
            tk.Label(row, text=label, font=(FF, 8), bg=C["surface"],
                     fg=C["text_muted"], width=10, anchor="w").pack(side="left")
            self._entry(row, textvariable=var, width=42).pack(side="left", fill="x", expand=True, ipady=4)
        ext_row = tk.Frame(opt_card, bg=C["surface"])
        ext_row.pack(fill="x", pady=(8, 0))
        self._btn(ext_row, "⟳ Auto-extract from JD",
                  self._auto_extract_job_details, style="ghost").pack(side="left")
        return opt_outer

    def _eval_context_label(self, parent):
        lbl = tk.Label(parent, text="", font=(FF, 8), bg=C["bg"])
        lbl._update = lambda: lbl.configure(
            text=("  ● Evaluation context available" if (self.assistant and self.assistant._last_evaluation)
                  else "  ○ No evaluation context"),
            fg=(C["success"] if (self.assistant and self.assistant._last_evaluation)
                else C["text_muted"]))
        return lbl

    def _word_count_label(self, parent, widget):
        lbl = tk.Label(parent, text="0 words · 0 chars",
                       font=(FF, 7), bg=C["bg"], fg=C["text_muted"])
        widget._count_label = lbl
        return lbl

    def _update_count(self, widget):
        lbl = getattr(widget, "_count_label", None)
        if lbl is None:
            return
        content = widget.get(1.0, "end-1c")
        words = len(content.split()) if content.strip() else 0
        lbl.configure(text=f"{words:,} words · {len(content.strip()):,} chars")

    # ═════════════════════════════════════════════════════════════════════════
    # Shared two-column page skeleton
    # ═════════════════════════════════════════════════════════════════════════

    def _two_col_page(self, parent, title, subtitle):
        """Create a page with a header + draggable 50/50 horizontal pane."""
        page  = tk.Frame(parent, bg=C["bg"])
        inner = tk.Frame(page, bg=C["bg"])
        inner.pack(fill="both", expand=True, padx=32, pady=24)
        self._section_header(inner, title, subtitle).pack(fill="x", pady=(0, 18))
        pane = tk.PanedWindow(inner, orient=tk.HORIZONTAL, bg=C["border_bright"],
                              sashwidth=4, sashrelief="flat", bd=0, relief="flat")
        pane.pack(fill="both", expand=True)
        left  = tk.Frame(pane, bg=C["bg"])
        right = tk.Frame(pane, bg=C["bg"])
        pane.add(left,  minsize=280, stretch="always")
        pane.add(right, minsize=280, stretch="always")
        return page, inner, left, right

    def _col_label(self, parent, text):
        tk.Label(parent, text=text, font=(FF, 7, "bold"),
                 bg=C["bg"], fg=C["text_muted"]).pack(anchor="w", pady=(0, 5))

    def _output_buttons(self, parent, widget, extra_btns=None):
        rb = tk.Frame(parent, bg=C["bg"])
        rb.pack(fill="x", pady=(6, 0))
        self._btn(rb, "Save",  lambda: self._save_text(widget),  style="secondary").pack(side="left", padx=(0, 6))
        self._btn(rb, "Copy",  lambda: self._copy_text(widget),  style="secondary").pack(side="left", padx=(0, 6))
        self._btn(rb, "Clear", lambda: self._clear_text(widget), style="ghost").pack(side="left")
        if extra_btns:
            for text, cmd, style in extra_btns:
                self._btn(rb, text, cmd, style=style).pack(side="left", padx=(6, 0))
        self._word_count_label(parent, widget).pack(anchor="e", pady=(4, 0))

    # ═════════════════════════════════════════════════════════════════════════
    # Pages
    # ═════════════════════════════════════════════════════════════════════════

    def _page_setup(self, parent):
        page  = tk.Frame(parent, bg=C["bg"])
        inner = tk.Frame(page, bg=C["bg"])
        inner.pack(fill="both", expand=True, padx=32, pady=24)
        self._section_header(inner, "Setup",
                             "Configure your CV path, model, and API credentials").pack(fill="x", pady=(0, 20))

        # CV card — multi-CV list
        cv_outer, cv_card = self._card(inner, "CV PROFILES")
        cv_outer.pack(fill="x", pady=(0, 12))

        tk.Label(cv_card, text="Saved CV versions — select one and click Use Selected to load it.",
                 font=(FF, 8), bg=C["surface"], fg=C["text_muted"]).pack(anchor="w", pady=(0, 6))

        cv_list_row = tk.Frame(cv_card, bg=C["surface"])
        cv_list_row.pack(fill="x")

        list_border = tk.Frame(cv_list_row, bg=C["border"], padx=1, pady=1)
        list_border.pack(side="left", fill="both", expand=True)
        self._cv_listbox = tk.Listbox(list_border, font=(FF, 9),
                                      bg=C["input_bg"], fg=C["text"],
                                      selectbackground=C["surface3"],
                                      selectforeground=C["accent_soft"],
                                      relief="flat", bd=0, height=4,
                                      activestyle="none")
        self._cv_listbox.pack(fill="both", expand=True, padx=4, pady=4)
        self._cv_listbox.bind("<Double-Button-1>", lambda e: self._cv_use_selected())

        cv_btns = tk.Frame(cv_list_row, bg=C["surface"])
        cv_btns.pack(side="left", padx=(8, 0), anchor="n")
        self._btn(cv_btns, "Add CV",       self._cv_add,         style="primary").pack(fill="x", pady=(0, 4))
        self._btn(cv_btns, "Use Selected", self._cv_use_selected, style="secondary").pack(fill="x", pady=(0, 4))
        self._btn(cv_btns, "Remove",       self._cv_remove,       style="danger").pack(fill="x")

        self._cv_active_lbl = tk.Label(cv_card, text="", font=(FF, 8),
                                       bg=C["surface"], fg=C["text_muted"], anchor="w")
        self._cv_active_lbl.pack(anchor="w", pady=(6, 0))

        self._cv_refresh_list()  # populate from self.cv_profiles

        # Provider & Model card
        api_outer, api_card = self._card(inner, "PROVIDER & MODEL")
        api_outer.pack(fill="x", pady=(0, 12))

        # Provider toggle
        prov_row = tk.Frame(api_card, bg=C["surface"])
        prov_row.pack(fill="x", pady=(0, 10))
        self._provider_var = tk.StringVar(value=self.provider)
        for label, val in [("OpenAI (cloud)", "openai"), ("Ollama (local, no API key)", "ollama")]:
            rb = tk.Radiobutton(prov_row, text=label, variable=self._provider_var,
                                value=val, font=(FF, 9), bg=C["surface"], fg=C["text"],
                                selectcolor=C["surface2"], activebackground=C["surface"],
                                relief="flat", command=self._on_provider_toggle)
            rb.pack(side="left", padx=(0, 16))

        # ── OpenAI section ────────────────────────────────────────────────
        self._openai_section = tk.Frame(api_card, bg=C["surface"])
        self._openai_section.pack(fill="x")

        tk.Label(self._openai_section, text="Select model", font=(FF, 8),
                 bg=C["surface"], fg=C["text_muted"]).pack(anchor="w", pady=(0, 6))
        model_row = tk.Frame(self._openai_section, bg=C["surface"])
        model_row.pack(anchor="w", pady=(0, 8))
        self._model_btns = {}
        for m in ["gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo", "gpt-5.2"]:
            b = tk.Button(model_row, text=m, relief="flat", bd=0, cursor="hand2",
                          font=(FF, 8), bg=C["surface3"], fg=C["text_dim"],
                          activebackground=C["accent_dim"], activeforeground=C["text"],
                          padx=12, pady=5,
                          command=lambda mv=m: self._select_model(mv))
            b.pack(side="left", padx=(0, 4))
            self._model_btns[m] = b
        self._select_model("gpt-4o")
        key_row = tk.Frame(self._openai_section, bg=C["surface"])
        key_row.pack(fill="x")
        tk.Label(key_row, text="API key", font=(FF, 8),
                 bg=C["surface"], fg=C["text_muted"]).pack(side="left", padx=(0, 10))
        self._api_status = tk.Label(key_row, text="Checking …", font=(FF, 8),
                                    bg=C["surface"], fg=C["warning"])
        self._api_status.pack(side="left")

        # ── Ollama section ────────────────────────────────────────────────
        self._ollama_section = tk.Frame(api_card, bg=C["surface"])
        # (not packed by default — shown when Ollama is selected)

        # URL row
        url_row = tk.Frame(self._ollama_section, bg=C["surface"])
        url_row.pack(fill="x", pady=(0, 8))
        tk.Label(url_row, text="Ollama URL", font=(FF, 8),
                 bg=C["surface"], fg=C["text_muted"], width=12, anchor="w").pack(side="left")
        url_entry = self._entry(url_row, textvariable=self._ollama_url_var, width=36)
        url_entry.pack(side="left", fill="x", expand=True, ipady=4, padx=(0, 8))
        self._btn(url_row, "Check", self._ollama_check_connection, style="ghost").pack(side="left")

        # Connection status
        self._ollama_status = tk.Label(self._ollama_section, text="● Not connected",
                                       font=(FF, 8), bg=C["surface"], fg=C["danger"], anchor="w")
        self._ollama_status.pack(anchor="w", pady=(0, 8))

        # Model list
        tk.Label(self._ollama_section, text="INSTALLED MODELS", font=(FF, 7, "bold"),
                 bg=C["surface"], fg=C["text_muted"]).pack(anchor="w", pady=(0, 4))
        ollama_list_row = tk.Frame(self._ollama_section, bg=C["surface"])
        ollama_list_row.pack(fill="x")

        list_border2 = tk.Frame(ollama_list_row, bg=C["border"], padx=1, pady=1)
        list_border2.pack(side="left", fill="both", expand=True)
        self._ollama_listbox = tk.Listbox(list_border2, font=(FM, 9),
                                          bg=C["input_bg"], fg=C["text"],
                                          selectbackground=C["surface3"],
                                          selectforeground=C["accent_soft"],
                                          relief="flat", bd=0, height=5,
                                          activestyle="none")
        self._ollama_listbox.pack(fill="both", expand=True, padx=4, pady=4)
        self._ollama_listbox.bind("<Double-Button-1>", lambda e: self._ollama_use_model())

        ollama_btns = tk.Frame(ollama_list_row, bg=C["surface"])
        ollama_btns.pack(side="left", padx=(8, 0), anchor="n")
        self._btn(ollama_btns, "Use Model", self._ollama_use_model, style="primary").pack(fill="x", pady=(0, 4))
        self._btn(ollama_btns, "Refresh",   self._ollama_refresh_models, style="secondary").pack(fill="x", pady=(0, 4))
        self._btn(ollama_btns, "Pull…",     self._ollama_pull_model, style="ghost").pack(fill="x")

        self._ollama_model_lbl = tk.Label(self._ollama_section, text="",
                                          font=(FF, 8), bg=C["surface"],
                                          fg=C["text_muted"], anchor="w")
        self._ollama_model_lbl.pack(anchor="w", pady=(6, 0))

        # Recommendations
        rec_outer, rec_card = self._card(self._ollama_section, padx=12, pady=8)
        rec_outer.pack(fill="x", pady=(10, 0))
        rec_card.configure(bg=C["surface2"])
        tk.Label(rec_card,
                 text="Recommended models for job applications:",
                 font=(FF, 7, "bold"), bg=C["surface2"], fg=C["text_dim"]).pack(anchor="w")
        for name, note in [("qwen2.5:7b",    "strong instruction following  (~5 GB RAM)"),
                            ("llama3.2:11b",  "balanced quality  (~8 GB RAM)"),
                            ("mistral-nemo",  "good for writing tasks  (~7 GB RAM)"),
                            ("llama3.2",      "fast, lighter  (~3 GB RAM)")]:
            tk.Label(rec_card, text=f"  {name:<20}  {note}",
                     font=(FM, 7), bg=C["surface2"], fg=C["text_muted"]).pack(anchor="w")
        tk.Label(rec_card,
                 text="Use 7B+ models for best results. Pull with: ollama pull <name>",
                 font=(FF, 7), bg=C["surface2"], fg=C["text_muted"]).pack(anchor="w", pady=(4, 0))

        # Apply saved provider state
        self._on_provider_toggle()

        # Action buttons
        btn_row = tk.Frame(inner, bg=C["bg"])
        btn_row.pack(fill="x", pady=14)
        self._btn(btn_row, "Initialize Assistant",
                  self._initialize_assistant, style="primary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Test Configuration",
                  self._test_configuration, style="secondary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Preview CV Text",
                  self._preview_cv_text, style="ghost").pack(side="left")

        # Log
        log_outer, log_card = self._card(inner, "LOG")
        log_outer.pack(fill="both", expand=True)
        log_card.configure(bg=C["surface"])
        self._setup_log, log_sb = self._textarea(log_card, height=14, mono=True)
        self._setup_log.configure(bg=C["sidebar_bg"], state="disabled")
        self._setup_log.pack(side="left", fill="both", expand=True)
        log_sb.pack(side="left", fill="y")
        return page

    def _page_evaluate(self, parent):
        page, _, left, right = self._two_col_page(
            parent, "Evaluate Fit",
            "Paste a job description — get a strategic analysis with salary estimate")

        self._col_label(left, "JOB DESCRIPTION")
        self.eval_job_text, _ = self._inset_area(left, height=22)
        self.eval_job_text._container.pack(fill="both", expand=True)
        lb = tk.Frame(left, bg=C["bg"])
        lb.pack(fill="x", pady=(10, 0))
        self._btn(lb, "⟶  Evaluate Fit",   self._evaluate_job_fit,                                          style="primary").pack(side="left", padx=(0, 8))
        self._btn(lb, "Paste",             lambda: self._paste_to(self.eval_job_text),                        style="ghost").pack(side="left", padx=(0, 6))
        self._btn(lb, "URL",               lambda: self._import_jd_from_url(self.eval_job_text),              style="ghost").pack(side="left", padx=(0, 6))
        self._btn(lb, "Clear",             lambda: self.eval_job_text.delete(1.0, "end"),                     style="ghost").pack(side="left")

        self._col_label(right, "EVALUATION RESULTS")
        self.eval_results, _ = self._inset_area(right, height=22)
        self.eval_results.configure(state="disabled")
        self.eval_results._container.pack(fill="both", expand=True)
        self._output_buttons(right, self.eval_results)
        return page

    def _page_generate(self, parent):
        page, _, left, right = self._two_col_page(
            parent, "Generate Materials",
            "CV summaries · cover letters · LinkedIn messages")

        self._col_label(left, "JOB DESCRIPTION")
        self.gen_job_text, _ = self._inset_area(left, height=13)
        self.gen_job_text._container.pack(fill="both", expand=True)
        self._optional_details_card(left, show_recruiter=True).pack(fill="x", pady=(10, 0))
        self._gen_eval_lbl = self._eval_context_label(left)
        self._gen_eval_lbl.pack(anchor="w", pady=(6, 0))
        lb = tk.Frame(left, bg=C["bg"])
        lb.pack(fill="x", pady=(8, 0))
        self._btn(lb, "CV Summary",       self._generate_cv_summary,              style="primary").pack(side="left", padx=(0, 6))
        self._btn(lb, "Cover Letter",     self._generate_cover_letter,            style="primary").pack(side="left", padx=(0, 6))
        self._btn(lb, "LinkedIn",         self._generate_linkedin_message,        style="outline").pack(side="left", padx=(0, 6))
        self._btn(lb, "ATS Check",        self._run_ats_check,                              style="outline").pack(side="left", padx=(0, 6))
        self._btn(lb, "Salary Negotiate", self._open_salary_negotiation,                    style="outline").pack(side="left", padx=(0, 6))
        self._btn(lb, "Paste",            lambda: self._paste_to(self.gen_job_text),         style="ghost").pack(side="left", padx=(0, 6))
        self._btn(lb, "URL",              lambda: self._import_jd_from_url(self.gen_job_text), style="ghost").pack(side="left")

        self._col_label(right, "GENERATED CONTENT")
        self.gen_output, _ = self._inset_area(right, height=24)
        self.gen_output.configure(state="disabled")
        self.gen_output._container.pack(fill="both", expand=True)
        self._output_buttons(right, self.gen_output)

        # Cover letter tone variations
        tone_row = tk.Frame(right, bg=C["bg"])
        tone_row.pack(fill="x", pady=(8, 0))
        tk.Label(tone_row, text="Cover letter tone:", font=(FF, 7),
                 bg=C["bg"], fg=C["text_muted"]).pack(side="left", padx=(0, 8))
        for label, tone in [("Hybrid", "hybrid"), ("Research-led", "research"), ("Engineering-led", "engineering")]:
            self._btn(tone_row, label,
                      lambda t=tone: self._generate_cover_letter(tone=t),
                      style="ghost").pack(side="left", padx=(0, 4))
        return page

    def _page_qa(self, parent):
        page, _, left, right = self._two_col_page(
            parent, "Application Q & A",
            "Answer specific questions from application forms or interviews")

        self._col_label(left, "JOB DESCRIPTION")
        self.qa_job_text, _ = self._inset_area(left, height=12)
        self.qa_job_text._container.pack(fill="both", expand=True)
        self._col_label(left, "YOUR QUESTION")
        self.qa_question, _ = self._inset_area(left, height=5)
        self.qa_question._container.pack(fill="x", pady=(5, 0))
        lb = tk.Frame(left, bg=C["bg"])
        lb.pack(fill="x", pady=(10, 0))
        self._btn(lb, "⟶  Answer",  self._answer_question,                          style="primary").pack(side="left", padx=(0, 8))
        self._btn(lb, "Paste",      lambda: self._paste_to(self.qa_job_text),        style="ghost").pack(side="left", padx=(0, 6))
        self._btn(lb, "URL",        lambda: self._import_jd_from_url(self.qa_job_text), style="ghost").pack(side="left")

        self._col_label(right, "AI ANSWER")
        self.qa_output, _ = self._inset_area(right, height=24)
        self.qa_output.configure(state="disabled")
        self.qa_output._container.pack(fill="both", expand=True)
        self._output_buttons(right, self.qa_output)
        return page

    def _page_package(self, parent):
        page, _, left, right = self._two_col_page(
            parent, "Complete Package",
            "Evaluation + CV summary + cover letter — streamed in sequence")

        self._col_label(left, "JOB DESCRIPTION")
        self.pkg_job_text, _ = self._inset_area(left, height=13)
        self.pkg_job_text._container.pack(fill="both", expand=True)
        self._optional_details_card(left).pack(fill="x", pady=(10, 0))
        lb = tk.Frame(left, bg=C["bg"])
        lb.pack(fill="x", pady=(12, 0))
        self._btn(lb, "⟶  Generate Complete Package",
                  self._generate_package, style="primary").pack(side="left", padx=(0, 8))
        self._btn(lb, "Paste",
                  lambda: self._paste_to(self.pkg_job_text),          style="ghost").pack(side="left", padx=(0, 6))
        self._btn(lb, "URL",
                  lambda: self._import_jd_from_url(self.pkg_job_text), style="ghost").pack(side="left")
        self._progress_lbl = tk.Label(left, text="", font=(FF, 8),
                                      bg=C["bg"], fg=C["accent"])
        self._progress_lbl.pack(anchor="w", pady=(8, 0))

        self._col_label(right, "GENERATED PACKAGE")
        self.pkg_output, _ = self._inset_area(right, height=24)
        self.pkg_output.configure(state="disabled")
        self.pkg_output._container.pack(fill="both", expand=True)
        self._output_buttons(right, self.pkg_output,
                             extra_btns=[("Save Package", lambda: self._save_text(self.pkg_output), "secondary")])
        return page

    def _page_interview(self, parent):
        page, _, left, right = self._two_col_page(
            parent, "Interview Prep",
            "Technical Q&A · STAR stories · smart questions to ask")

        self._col_label(left, "JOB DESCRIPTION")
        self.interview_job_text, _ = self._inset_area(left, height=22)
        self.interview_job_text._container.pack(fill="both", expand=True)
        self._int_eval_lbl = self._eval_context_label(left)
        self._int_eval_lbl.pack(anchor="w", pady=(6, 0))
        lb = tk.Frame(left, bg=C["bg"])
        lb.pack(fill="x", pady=(8, 0))
        self._btn(lb, "⟶  Generate Interview Prep",
                  self._generate_interview_prep,                              style="primary").pack(side="left", padx=(0, 8))
        self._btn(lb, "Paste",
                  lambda: self._paste_to(self.interview_job_text),                 style="ghost").pack(side="left", padx=(0, 6))
        self._btn(lb, "URL",
                  lambda: self._import_jd_from_url(self.interview_job_text),       style="ghost").pack(side="left", padx=(0, 6))
        self._btn(lb, "Clear",
                  lambda: self.interview_job_text.delete(1.0, "end"),              style="ghost").pack(side="left")

        # Follow-up email section (below the button row on the left)
        tk.Frame(left, bg=C["border"], height=1).pack(fill="x", pady=(12, 0))
        self._col_label(left, "FOLLOW-UP / THANK-YOU EMAIL")

        fu_name_row = tk.Frame(left, bg=C["bg"])
        fu_name_row.pack(fill="x", pady=(0, 4))
        tk.Label(fu_name_row, text="Interviewer name", font=(FF, 8),
                 bg=C["bg"], fg=C["text_muted"], width=16, anchor="w").pack(side="left")
        self._entry(fu_name_row, textvariable=self._followup_name_var,
                    width=30).pack(side="left", fill="x", expand=True, ipady=4)

        tk.Label(left, text="Interview notes  (optional)", font=(FF, 8),
                 bg=C["bg"], fg=C["text_muted"]).pack(anchor="w", pady=(4, 2))
        fu_notes_border = tk.Frame(left, bg=C["border_bright"], padx=1, pady=1)
        fu_notes_border.pack(fill="x")
        self.followup_notes_text = tk.Text(
            fu_notes_border, height=3, font=(FF, 9),
            bg=C["input_bg"], fg=C["text"], insertbackground=C["accent_soft"],
            relief="flat", bd=0, wrap="word", padx=8, pady=6)
        self.followup_notes_text.pack(fill="x")
        self.followup_notes_text.bind("<FocusIn>",  lambda e: fu_notes_border.configure(bg=C["accent"]))
        self.followup_notes_text.bind("<FocusOut>", lambda e: fu_notes_border.configure(bg=C["border_bright"]))

        fu_btn_row = tk.Frame(left, bg=C["bg"])
        fu_btn_row.pack(fill="x", pady=(6, 0))
        self._btn(fu_btn_row, "⟶  Generate Follow-Up Email",
                  self._generate_followup_email, style="primary").pack(side="left")

        self._col_label(right, "INTERVIEW PREPARATION")
        self.interview_output, _ = self._inset_area(right, height=22)
        self.interview_output.configure(state="disabled")
        self.interview_output._container.pack(fill="both", expand=True)
        self._output_buttons(right, self.interview_output,
                             extra_btns=[("⚡ Quiz Mode", self._open_quiz_mode, "outline")])
        return page

    def _page_history(self, parent):
        page  = tk.Frame(parent, bg=C["bg"])
        inner = tk.Frame(page, bg=C["bg"])
        inner.pack(fill="both", expand=True, padx=32, pady=24)
        self._section_header(inner, "History",
                             "All generated results — searchable, filterable, pipeline-tracked").pack(fill="x", pady=(0, 18))
        pane = tk.PanedWindow(inner, orient=tk.HORIZONTAL, bg=C["border_bright"],
                              sashwidth=4, sashrelief="flat", bd=0, relief="flat")
        pane.pack(fill="both", expand=True)

        # Left panel — list
        left = tk.Frame(pane, bg=C["bg"])
        pane.add(left, minsize=240, width=340, stretch="never")

        # Search
        tk.Label(left, text="SEARCH", font=(FF, 7, "bold"),
                 bg=C["bg"], fg=C["text_muted"]).pack(anchor="w", pady=(0, 4))
        search_ent = self._entry(left, textvariable=self._hist_search_var, width=20)
        search_ent.pack(fill="x", pady=(0, 8), ipady=5)
        self._hist_search_var.trace_add("write", lambda *_: self._refresh_hist_list())

        # Filter buttons
        filter_row = tk.Frame(left, bg=C["bg"])
        filter_row.pack(fill="x", pady=(0, 8))
        self._hist_filter_btns = {}
        for label, fstatus in [("All", ""), ("Applied", "Applied"), ("Intv", "Interview"),
                                ("Offer", "Offer"), ("Rejctd", "Rejected")]:
            btn = tk.Button(filter_row, text=label, relief="flat", bd=0, cursor="hand2",
                            font=(FF, 8), padx=8, pady=3,
                            bg=C["accent"] if fstatus == "" else C["surface2"],
                            fg=C["text_inv"] if fstatus == "" else C["text_dim"],
                            activebackground=C["accent_dim"], activeforeground=C["text"],
                            command=lambda s=fstatus: self._set_hist_filter(s))
            btn.pack(side="left", padx=(0, 3))
            self._hist_filter_btns[fstatus] = btn

        # Entry list
        tk.Label(left, text="ENTRIES", font=(FF, 7, "bold"),
                 bg=C["bg"], fg=C["text_muted"]).pack(anchor="w", pady=(0, 4))
        list_border = tk.Frame(left, bg=C["border"], padx=1, pady=1)
        list_border.pack(fill="both", expand=True)
        list_inner = tk.Frame(list_border, bg=C["surface"])
        list_inner.pack(fill="both", expand=True)
        self._hist_list = tk.Listbox(list_inner, font=(FM, 8),
                                     bg=C["surface"], fg=C["text"],
                                     selectbackground=C["surface3"], selectforeground=C["accent_soft"],
                                     relief="flat", bd=0, activestyle="none")
        hist_sb = tk.Scrollbar(list_inner, command=self._hist_list.yview,
                               bg=C["surface2"], troughcolor=C["sidebar_bg"], relief="flat")
        self._hist_list.configure(yscrollcommand=hist_sb.set)
        self._hist_list.pack(side="left", fill="both", expand=True)
        hist_sb.pack(side="left", fill="y")
        self._hist_list.bind("<<ListboxSelect>>", self._on_hist_select)
        self._hist_count_lbl = tk.Label(left, text="", font=(FF, 7),
                                        bg=C["bg"], fg=C["text_muted"])
        self._hist_count_lbl.pack(anchor="w", pady=(5, 0))
        hb = tk.Frame(left, bg=C["bg"])
        hb.pack(fill="x", pady=(8, 0))
        self._btn(hb, "Delete", self._delete_hist_entry, style="danger").pack(side="left", padx=(0, 6))
        self._btn(hb, "Clear All", self._clear_history,  style="ghost").pack(side="left")

        hb2 = tk.Frame(left, bg=C["bg"])
        hb2.pack(fill="x", pady=(5, 0))
        self._btn(hb2, "Export CSV",  self._history_export_csv,  style="secondary").pack(side="left", padx=(0, 6))
        self._btn(hb2, "Export JSON", self._history_export_json, style="secondary").pack(side="left")

        # Right panel — viewer
        right = tk.Frame(pane, bg=C["bg"])
        pane.add(right, minsize=300, stretch="always")
        self._hist_meta = tk.Label(right, text="Select an entry to view its content",
                                   font=(FF, 8), bg=C["bg"], fg=C["text_muted"], anchor="w")
        self._hist_meta.pack(anchor="w", pady=(0, 6))
        # Status row
        status_row = tk.Frame(right, bg=C["bg"])
        status_row.pack(fill="x", pady=(0, 8))
        tk.Label(status_row, text="Status:", font=(FF, 8),
                 bg=C["bg"], fg=C["text_muted"]).pack(side="left", padx=(0, 8))
        for label, style in [("Applied", "secondary"), ("Interview", "warning"),
                              ("Offer", "success"), ("Rejected", "danger"), ("Clear", "ghost")]:
            self._btn(status_row, label,
                      lambda s=label: self._set_hist_status(s),
                      style=style).pack(side="left", padx=(0, 4))

        self.hist_content, _ = self._inset_area(right, height=16)
        self.hist_content.configure(state="disabled")
        self.hist_content._container.pack(fill="both", expand=True)

        # Personal notes field — editable, auto-saves on focus-out
        tk.Label(right, text="PERSONAL NOTES", font=(FF, 7, "bold"),
                 bg=C["bg"], fg=C["text_muted"]).pack(anchor="w", pady=(10, 3))
        notes_border = tk.Frame(right, bg=C["border_bright"], padx=1, pady=1)
        notes_border.pack(fill="x")
        self._hist_notes = tk.Text(notes_border, height=3, font=(FF, 9),
                                   bg=C["input_bg"], fg=C["text"],
                                   insertbackground=C["accent_soft"],
                                   relief="flat", bd=0, wrap="word",
                                   padx=8, pady=6, undo=True)
        self._hist_notes.pack(fill="x")
        self._hist_notes.bind("<FocusIn>",  lambda e: notes_border.configure(bg=C["accent"]))
        self._hist_notes.bind("<FocusOut>", lambda e: (notes_border.configure(bg=C["border_bright"]),
                                                        self._save_hist_notes()))
        self._hist_notes.insert(1.0, "Select an entry to add notes…")
        self._hist_notes.configure(state="disabled", fg=C["text_muted"])

        rb = tk.Frame(right, bg=C["bg"])
        rb.pack(fill="x", pady=(8, 0))
        self._btn(rb, "Save Entry",       lambda: self._save_text(self.hist_content),    style="secondary").pack(side="left", padx=(0, 6))
        self._btn(rb, "Copy",             lambda: self._copy_text(self.hist_content),    style="secondary").pack(side="left", padx=(0, 6))
        self._btn(rb, "+ Add to Tracker", self._add_to_tracker_from_history,             style="ghost").pack(side="left", padx=(0, 6))
        self._hist_rejection_btn = self._btn(rb, "Analyze Rejection",
                                             self._analyze_rejection_from_history,
                                             style="danger")
        self._hist_rejection_btn.pack_forget()   # hidden until a Rejected entry is selected
        self._refresh_hist_list()
        return page

    def _page_profile(self, parent):
        page  = tk.Frame(parent, bg=C["bg"])
        inner = tk.Frame(page, bg=C["bg"])
        inner.pack(fill="both", expand=True, padx=32, pady=24)
        self._section_header(inner, "Profile Editor",
                             "Edit your profile — changes apply immediately on save").pack(fill="x", pady=(0, 16))

        # Style the notebook to match the dark theme
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Profile.TNotebook",
                        background=C["bg"],
                        borderwidth=0,
                        tabmargins=[0, 0, 0, 0])
        style.configure("Profile.TNotebook.Tab",
                        background=C["surface"],
                        foreground=C["text_dim"],
                        padding=[14, 6],
                        font=(FF, 9),
                        borderwidth=0,
                        focuscolor=C["surface"])
        style.map("Profile.TNotebook.Tab",
                  background=[("selected", C["surface2"])],
                  foreground=[("selected", C["accent_soft"])],
                  expand=[("selected", [0, 0, 0, 0])])

        # Button bar at bottom first (so editors never push it off screen)
        btn_row = tk.Frame(inner, bg=C["bg"])
        btn_row.pack(side="bottom", fill="x", pady=(8, 0))
        self._btn(btn_row, "Save Both",        self._save_profile,        style="primary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Reload from Disk", self._load_profile_editor, style="secondary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Setup Wizard",     self._open_profile_wizard, style="ghost").pack(side="left")
        self._profile_status = tk.Label(btn_row, text="", font=(FF, 8),
                                        bg=C["bg"], fg=C["text_muted"])
        self._profile_status.pack(side="right")

        # Notebook with two tabs
        nb = ttk.Notebook(inner, style="Profile.TNotebook")
        nb.pack(fill="both", expand=True)

        # ── Tab 1: Personal Profile ──────────────────────────────────────────
        tab_personal = tk.Frame(nb, bg=C["bg"])
        nb.add(tab_personal, text="  Personal Profile  ")

        personal_info_outer, personal_info_card = self._card(tab_personal, padx=14, pady=10)
        personal_info_outer.pack(fill="x", pady=(12, 6))
        self._profile_info_outer = personal_info_outer
        tk.Label(personal_info_card,
                 text="Your personal info — name, skills, career goals, salary benchmarks, CV highlights. "
                      "Fill in all [YOUR ...] placeholders.",
                 font=(FF, 8), bg=C["surface"], fg=C["text_dim"],
                 wraplength=900, justify="left").pack(anchor="w")

        # Placeholder warning banner for personal tab
        self._profile_warn = tk.Label(tab_personal, text="", font=(FF, 8),
                                      bg=C["warning"], fg=C["text_inv"],
                                      wraplength=900, justify="left",
                                      padx=10, pady=6, anchor="w")

        personal_border = tk.Frame(tab_personal, bg=C["border_bright"], padx=1, pady=1)
        personal_border.pack(fill="both", expand=True, pady=(0, 4))
        personal_inner = tk.Frame(personal_border, bg=C["input_bg"])
        personal_inner.pack(fill="both", expand=True)
        self.profile_personal_editor, personal_sb = self._textarea(personal_inner, height=30, mono=True)
        self.profile_personal_editor.pack(side="left", fill="both", expand=True)
        personal_sb.pack(side="left", fill="y")
        self.profile_personal_editor.bind("<FocusIn>",  lambda e: personal_border.configure(bg=C["accent"]))
        self.profile_personal_editor.bind("<FocusOut>", lambda e: personal_border.configure(bg=C["border_bright"]))

        # Keep legacy attribute pointing to personal editor for backward compat
        self.profile_editor = self.profile_personal_editor

        # ── Tab 2: AI Instructions ───────────────────────────────────────────
        tab_instr = tk.Frame(nb, bg=C["bg"])
        nb.add(tab_instr, text="  AI Instructions  ")

        instr_info_outer, instr_info_card = self._card(tab_instr, padx=14, pady=10)
        instr_info_outer.pack(fill="x", pady=(12, 6))
        tk.Label(instr_info_card,
                 text="AI behavior — writing tone, output rules, how to handle job descriptions. "
                      "Modify only if you want to change how the assistant behaves.",
                 font=(FF, 8), bg=C["surface"], fg=C["text_dim"],
                 wraplength=900, justify="left").pack(anchor="w")

        # Warning label for instructions tab (separate from personal tab warn)
        self._profile_instr_warn = tk.Label(tab_instr, text="", font=(FF, 8),
                                            bg=C["warning"], fg=C["text_inv"],
                                            wraplength=900, justify="left",
                                            padx=10, pady=6, anchor="w")

        instr_border = tk.Frame(tab_instr, bg=C["border_bright"], padx=1, pady=1)
        instr_border.pack(fill="both", expand=True, pady=(0, 4))
        instr_inner = tk.Frame(instr_border, bg=C["input_bg"])
        instr_inner.pack(fill="both", expand=True)
        self.profile_instr_editor, instr_sb = self._textarea(instr_inner, height=30, mono=True)
        self.profile_instr_editor.pack(side="left", fill="both", expand=True)
        instr_sb.pack(side="left", fill="y")
        self.profile_instr_editor.bind("<FocusIn>",  lambda e: instr_border.configure(bg=C["accent"]))
        self.profile_instr_editor.bind("<FocusOut>", lambda e: instr_border.configure(bg=C["border_bright"]))

        self._load_profile_editor()
        return page

    def _page_stats(self, parent):
        page  = tk.Frame(parent, bg=C["bg"])
        inner = tk.Frame(page, bg=C["bg"])
        inner.pack(fill="both", expand=True, padx=32, pady=24)
        hdr_row = tk.Frame(inner, bg=C["bg"])
        hdr_row.pack(fill="x", pady=(0, 18))
        self._section_header(hdr_row, "Statistics",
                             "All-time pipeline health, material usage, and cost").pack(
            side="left", fill="x", expand=True)
        self._btn(hdr_row, "⟳  Refresh", self._refresh_stats, style="ghost").pack(
            side="right", padx=(8, 0))

        # We'll build the stats content into a scrollable canvas
        canvas = tk.Canvas(inner, bg=C["bg"], highlightthickness=0)
        sb = tk.Scrollbar(inner, orient="vertical", command=canvas.yview,
                          bg=C["surface2"], troughcolor=C["sidebar_bg"], relief="flat")
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        canvas.configure(yscrollcommand=sb.set)
        canvas.bind("<MouseWheel>",
                    lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        self._stats_frame = tk.Frame(canvas, bg=C["bg"])
        win = canvas.create_window((0, 0), window=self._stats_frame, anchor="nw")
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(win, width=e.width))
        self._stats_frame.bind("<Configure>",
                               lambda e: canvas.configure(
                                   scrollregion=canvas.bbox("all")))

        self._refresh_stats()
        return page

    def _refresh_stats(self):
        if not hasattr(self, "_stats_frame"):
            return
        for w in self._stats_frame.winfo_children():
            w.destroy()
        f = self._stats_frame
        pad = {"padx": 0, "pady": (0, 14)}

        # ── Helper ───────────────────────────────────────────────────────────
        def stat_pill(parent, label, value, fg_color, subtitle=None):
            p = tk.Frame(parent, bg=C["surface2"], padx=14, pady=8)
            p.pack(side="left", padx=(0, 8), pady=(0, 4))
            tk.Label(p, text=value, font=(FF, 14, "bold"),
                     bg=C["surface2"], fg=fg_color).pack(anchor="w")
            tk.Label(p, text=label, font=(FF, 7),
                     bg=C["surface2"], fg=C["text_muted"]).pack(anchor="w")
            if subtitle:
                tk.Label(p, text=subtitle, font=(FF, 6),
                         bg=C["surface2"], fg=C["text_muted"]).pack(anchor="w")

        # ── Pipeline (current status) ─────────────────────────────────────────
        _, card = self._card(f, "CURRENT PIPELINE  (by current status)", padx=16, pady=14)
        card.master.pack(fill="x", **pad)
        card.configure(bg=C["surface"])

        total_apps = len(self.applications)
        counts = {s: 0 for s in STATUSES}
        for a in self.applications:
            s = a.get("status", "Applied")
            if s in counts:
                counts[s] += 1

        # Status grid
        grid = tk.Frame(card, bg=C["surface"])
        grid.pack(fill="x", pady=(0, 10))
        for i, status in enumerate(STATUSES):
            _, fg = STATUS_ROW_COLORS.get(status, ("#181b26", C["text_muted"]))
            col = i % 5
            row = i // 5
            cell = tk.Frame(grid, bg=C["surface2"], padx=12, pady=8)
            cell.grid(row=row, column=col, padx=(0, 6), pady=(0, 6), sticky="ew")
            grid.columnconfigure(col, weight=1)
            tk.Label(cell, text=str(counts[status]), font=(FF, 18, "bold"),
                     bg=C["surface2"], fg=fg).pack(anchor="w")
            tk.Label(cell, text=status, font=(FF, 7),
                     bg=C["surface2"], fg=C["text_muted"]).pack(anchor="w")

        # ── Funnel (using peak_stage — correct even after rejection) ──────────
        _, card_f = self._card(f, "CONVERSION FUNNEL  (by peak stage reached)", padx=16, pady=14)
        card_f.master.pack(fill="x", **pad)
        card_f.configure(bg=C["surface"])

        tk.Label(card_f,
                 text="Peak stage tracks the furthest point each application reached — "
                      "so rejections after interview still count as interviewed.",
                 font=(FF, 7), bg=C["surface"], fg=C["text_muted"],
                 wraplength=700, justify="left").pack(anchor="w", pady=(0, 10))

        # Count by peak_stage (fall back to current status for entries without peak_stage)
        peak_counts = {s: 0 for s in STAGE_ORDER}
        for a in self.applications:
            peak = a.get("peak_stage") or (
                a.get("status") if a.get("status") not in TERMINAL_STATUSES else "Applied")
            if peak in peak_counts:
                peak_counts[peak] += 1

        funnel_row = tk.Frame(card_f, bg=C["surface"])
        funnel_row.pack(fill="x", pady=(0, 10))
        stage_order_list = sorted(STAGE_ORDER.items(), key=lambda x: x[1])
        for stage, _ in stage_order_list:
            _, fg = STATUS_ROW_COLORS.get(stage, ("#181b26", C["text_muted"]))
            cell = tk.Frame(funnel_row, bg=C["surface2"], padx=12, pady=8)
            cell.pack(side="left", padx=(0, 6))
            tk.Label(cell, text=str(peak_counts[stage]), font=(FF, 18, "bold"),
                     bg=C["surface2"], fg=fg).pack(anchor="w")
            tk.Label(cell, text=stage, font=(FF, 7),
                     bg=C["surface2"], fg=C["text_muted"]).pack(anchor="w")

        # Funnel rates (based on peak_stage — the honest numbers)
        n_total      = total_apps
        n_responded  = sum(peak_counts[s] for s in
                           ["Phone Screen", "Interview", "Final Round", "Offer", "Accepted"])
        n_interviewed = sum(peak_counts[s] for s in
                            ["Interview", "Final Round", "Offer", "Accepted"])
        n_offers     = peak_counts.get("Offer", 0) + peak_counts.get("Accepted", 0)

        rates = tk.Frame(card_f, bg=C["surface"])
        rates.pack(fill="x")
        stat_pill(rates, "Total Applied", str(n_total), C["text"])
        resp_pct = f"{n_responded/n_total*100:.0f}%" if n_total else "—"
        stat_pill(rates, "Response Rate",
                  resp_pct, C["warning"],
                  subtitle=f"{n_responded} got to Phone Screen+")
        int_pct = f"{n_interviewed/n_total*100:.0f}%" if n_total else "—"
        stat_pill(rates, "Interview Rate",
                  int_pct, C["accent_soft"],
                  subtitle=f"{n_interviewed} reached Interview+")
        offer_pct = f"{n_offers/n_total*100:.0f}%" if n_total else "—"
        stat_pill(rates, "Offer Rate",
                  offer_pct, C["success"],
                  subtitle=f"{n_offers} received an offer")

        # ── Generated Materials ───────────────────────────────────────────────
        _, card2 = self._card(f, "GENERATED MATERIALS", padx=16, pady=14)
        card2.master.pack(fill="x", **pad)
        card2.configure(bg=C["surface"])

        type_counts: dict[str, int] = {}
        total_hist_tokens = 0
        total_hist_cost   = 0.0
        for h in self.history:
            t = h.get("type", "Unknown")
            type_counts[t] = type_counts.get(t, 0) + 1
            toks = h.get("tokens", 0) or 0
            total_hist_tokens += toks
            rate = MODEL_COSTS.get(h.get("model", "gpt-4o"), 0.010)
            total_hist_cost += (toks / 1000) * rate

        mat_grid = tk.Frame(card2, bg=C["surface"])
        mat_grid.pack(fill="x", pady=(0, 10))
        sorted_types = sorted(type_counts.items(), key=lambda x: -x[1])
        for i, (t, c) in enumerate(sorted_types):
            col = i % 4
            row = i // 4
            cell = tk.Frame(mat_grid, bg=C["surface2"], padx=10, pady=6)
            cell.grid(row=row, column=col, padx=(0, 6), pady=(0, 6), sticky="ew")
            mat_grid.columnconfigure(col, weight=1)
            tk.Label(cell, text=str(c), font=(FF, 14, "bold"),
                     bg=C["surface2"], fg=C["accent_soft"]).pack(anchor="w")
            tk.Label(cell, text=t, font=(FF, 7),
                     bg=C["surface2"], fg=C["text_muted"]).pack(anchor="w")

        tok_row = tk.Frame(card2, bg=C["surface"])
        tok_row.pack(fill="x")
        cost_display = "Local (free)" if self.provider == "ollama" else f"${total_hist_cost:.2f}"
        stat_pill(tok_row, "Total Entries", str(len(self.history)), C["text"])
        stat_pill(tok_row, "All-time Tokens", f"{total_hist_tokens:,}", C["text_dim"])
        stat_pill(tok_row, "Estimated Cost", cost_display, C["warning"])

        # ── Current Session ───────────────────────────────────────────────────
        _, card3 = self._card(f, "CURRENT SESSION", padx=16, pady=14)
        card3.master.pack(fill="x", **pad)
        card3.configure(bg=C["surface"])
        sess_row = tk.Frame(card3, bg=C["surface"])
        sess_row.pack(fill="x")
        stat_pill(sess_row, "Session Tokens", f"{self._session_tokens:,}", C["accent_soft"])
        stat_pill(sess_row, "Session Cost",   f"${self._session_cost:.3f}", C["warning"])
        active_cv = self.active_cv_name or Path(self.cv_path.get()).name
        stat_pill(sess_row, "Active CV", active_cv, C["text_dim"])

    def _page_tracker(self, parent):
        page  = tk.Frame(parent, bg=C["bg"])
        inner = tk.Frame(page, bg=C["bg"])
        inner.pack(fill="both", expand=True, padx=32, pady=24)
        self._section_header(inner, "Application Tracker",
                             "Full pipeline view — track every application from first look to offer").pack(fill="x", pady=(0, 14))
        self._configure_tracker_style()

        # Toolbar
        tb = tk.Frame(inner, bg=C["bg"])
        tb.pack(fill="x", pady=(0, 8))
        self._btn(tb, "+ New Application", self._tracker_new_entry,  style="primary").pack(side="left", padx=(0, 8))
        self._btn(tb, "Export CSV",        self._tracker_export_csv, style="secondary").pack(side="left", padx=(0, 8))
        self._btn(tb, "Import CSV/XLSX",   self._tracker_import_file, style="secondary").pack(side="left")

        # Pipeline stats bar — click a status to filter; click again to clear
        stats_bar = tk.Frame(inner, bg=C["bg"])
        stats_bar.pack(fill="x", pady=(0, 10))
        self._stats_count_labels = {}
        for status, (_, fg) in STATUS_ROW_COLORS.items():
            cell = tk.Frame(stats_bar, bg=C["bg"], cursor="hand2")
            cell.pack(side="left", padx=(0, 2))
            dot = tk.Label(cell, text="●", font=(FF, 8), bg=C["bg"], fg=fg)
            dot.pack(side="left")
            name_lbl = tk.Label(cell, text=status, font=(FF, 7), bg=C["bg"],
                                fg=C["text_muted"])
            name_lbl.pack(side="left", padx=(1, 2))
            cnt_lbl = tk.Label(cell, text="0", font=(FF, 7, "bold"), bg=C["bg"],
                               fg=fg)
            cnt_lbl.pack(side="left", padx=(0, 8))
            self._stats_count_labels[status] = (cell, cnt_lbl, name_lbl, dot)
            for widget in (cell, dot, name_lbl, cnt_lbl):
                widget.bind("<Button-1>",
                            lambda e, s=status: self._tracker_filter_by_status(s))

        # Vertical pane: table top / edit form bottom
        vpane = tk.PanedWindow(inner, orient=tk.VERTICAL, bg=C["border_bright"],
                               sashwidth=4, sashrelief="flat", bd=0, relief="flat")
        vpane.pack(fill="both", expand=True)

        # Table pane
        table_frame = tk.Frame(vpane, bg=C["bg"])
        vpane.add(table_frame, minsize=120, stretch="always")
        tree_border = tk.Frame(table_frame, bg=C["border"], padx=1, pady=1)
        tree_border.pack(fill="both", expand=True)
        tree_inner = tk.Frame(tree_border, bg=C["surface"])
        tree_inner.pack(fill="both", expand=True)
        vsb = tk.Scrollbar(tree_inner, orient="vertical",
                           bg=C["surface2"], troughcolor=C["sidebar_bg"], relief="flat")
        vsb.pack(side="right", fill="y")
        xsb = tk.Scrollbar(tree_inner, orient="horizontal",
                           bg=C["surface2"], troughcolor=C["sidebar_bg"], relief="flat")
        xsb.pack(side="bottom", fill="x")
        self._tracker_tree = ttk.Treeview(tree_inner, style="Tracker.Treeview",
                                          columns=[c[0] for c in TRACKER_COLS],
                                          show="headings", selectmode="browse",
                                          yscrollcommand=vsb.set, xscrollcommand=xsb.set)
        vsb.configure(command=self._tracker_tree.yview)
        xsb.configure(command=self._tracker_tree.xview)
        for col_id, col_hdr, col_w in TRACKER_COLS:
            self._tracker_tree.heading(col_id, text=col_hdr, anchor="w",
                                       command=lambda c=col_id: self._tracker_sort_by(c))
            self._tracker_tree.column(col_id, width=col_w, minwidth=50, anchor="w", stretch=False)
        for status, (bg, fg) in STATUS_ROW_COLORS.items():
            self._tracker_tree.tag_configure(status.replace(" ", "_"), background=bg, foreground=fg)
        self._tracker_tree.pack(side="left", fill="both", expand=True)
        self._tracker_tree.bind("<<TreeviewSelect>>", self._on_tracker_select)
        self._tracker_tree.bind("<MouseWheel>",
            lambda e: self._tracker_tree.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        # Edit / Add entry pane
        edit_frame = tk.Frame(vpane, bg=C["bg"])
        vpane.add(edit_frame, minsize=260, stretch="never")
        edit_outer, edit_card = self._card(edit_frame, "EDIT / ADD ENTRY", padx=14, pady=12)
        edit_outer.pack(fill="both", expand=True, pady=(6, 0))
        edit_card.configure(bg=C["surface"])
        form = tk.Frame(edit_card, bg=C["surface"])
        form.pack(fill="x", pady=(0, 8))
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        def lbl(text, row, col):
            tk.Label(form, text=text, font=(FF, 8), bg=C["surface"],
                     fg=C["text_muted"], width=12, anchor="e").grid(
                row=row, column=col, padx=(0, 6), pady=3, sticky="e")

        def ent(var, row, col):
            e = tk.Entry(form, textvariable=var, font=(FF, 9),
                         bg=C["input_bg"], fg=C["text"],
                         insertbackground=C["accent_soft"], relief="flat", bd=0,
                         highlightthickness=1,
                         highlightcolor=C["accent"],
                         highlightbackground=C["border"])
            e.grid(row=row, column=col, padx=(0, 14), pady=3, sticky="ew", ipady=5)
            return e

        lbl("Company",     0, 0); ent(self._trk_company,   0, 1)
        lbl("Position",    0, 2); ent(self._trk_position,  0, 3)
        lbl("Location",    1, 0); ent(self._trk_location,  1, 1)
        lbl("Date",        1, 2); ent(self._trk_date,       1, 3)
        lbl("Source",      2, 0); ent(self._trk_source,    2, 1)
        lbl("Salary (JD)", 2, 2); ent(self._trk_salary_jd, 2, 3)
        lbl("Requested",   3, 0); ent(self._trk_salary_req, 3, 1)
        lbl("Status",      3, 2)
        status_menu = tk.OptionMenu(form, self._trk_status_var, *STATUSES)
        status_menu.configure(font=(FF, 9), bg=C["input_bg"], fg=C["text"],
                              relief="flat", bd=0, highlightthickness=0,
                              activebackground=C["accent_dim"], activeforeground=C["text"])
        status_menu["menu"].configure(font=(FF, 9), bg=C["input_bg"], fg=C["text"],
                                      relief="flat", bd=0,
                                      activebackground=C["accent_dim"], activeforeground=C["text"])
        status_menu.grid(row=3, column=3, padx=(0, 14), pady=3, sticky="ew")
        lbl("Interview Date", 4, 0); ent(self._trk_date_interview, 4, 1)
        lbl("Decision Date",  4, 2); ent(self._trk_date_decision,  4, 3)
        # Peak stage row — read-only indicator
        lbl("Peak Stage", 5, 0)
        self._trk_peak_lbl = tk.Label(form, textvariable=self._trk_peak_stage,
                                      font=(FF, 8), bg=C["surface"],
                                      fg=C["text_muted"], anchor="w")
        self._trk_peak_lbl.grid(row=5, column=1, padx=(0, 14), pady=3, sticky="w")
        lbl("Notes", 6, 0)
        nb = tk.Frame(form, bg=C["border"], padx=1, pady=1)
        nb.grid(row=6, column=1, columnspan=3, padx=(0, 14), pady=3, sticky="ew")
        self._trk_notes = tk.Text(nb, height=3, font=(FF, 9),
                                   bg=C["input_bg"], fg=C["text"],
                                   insertbackground=C["accent_soft"], relief="flat", bd=0,
                                   wrap="word", padx=8, pady=6)
        self._trk_notes.pack(fill="x")
        self._trk_notes.bind("<FocusIn>",  lambda e: nb.configure(bg=C["accent"]))
        self._trk_notes.bind("<FocusOut>", lambda e: nb.configure(bg=C["border"]))

        btn_row = tk.Frame(edit_card, bg=C["surface"])
        btn_row.pack(fill="x")
        self._btn(btn_row, "Save Entry",   self._tracker_save_entry,   style="primary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Delete Entry", self._tracker_delete_entry, style="danger").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Clear Form",   self._tracker_new_entry,    style="ghost").pack(side="left")
        self._trk_status_lbl = tk.Label(btn_row,
                                         text="Select a row to edit, or click + New Application",
                                         font=(FF, 8), bg=C["surface"], fg=C["text_muted"])
        self._trk_status_lbl.pack(side="right", padx=(0, 4))
        self._refresh_tracker_table()
        return page

    # ═════════════════════════════════════════════════════════════════════════
    # Streaming helpers
    # ═════════════════════════════════════════════════════════════════════════

    def _stream_into(self, widget, prefix=""):
        widget.configure(state="normal")
        widget.delete(1.0, "end")
        if prefix:
            widget.insert("end", prefix)
        widget.configure(state="disabled")
        self._update_count(widget)

        def _callback(chunk):
            def _do():
                widget.configure(state="normal")
                widget.insert("end", chunk)
                widget.see("end")
                widget.configure(state="disabled")
                self._update_count(widget)
            self.root.after(0, _do)
        return _callback

    def _append_to(self, widget, text):
        def _do():
            widget.configure(state="normal")
            widget.insert("end", text)
            widget.see("end")
            widget.configure(state="disabled")
            self._update_count(widget)
        self.root.after(0, _do)

    def _make_append_callback(self, widget):
        def _callback(chunk):
            def _do():
                widget.configure(state="normal")
                widget.insert("end", chunk)
                widget.see("end")
                widget.configure(state="disabled")
                self._update_count(widget)
            self.root.after(0, _do)
        return _callback

    # ═════════════════════════════════════════════════════════════════════════
    # Setup actions
    # ═════════════════════════════════════════════════════════════════════════

    # ─────────────────────────────────────────────────────────────────────────
    # Provider / Ollama helpers
    # ─────────────────────────────────────────────────────────────────────────

    def _on_provider_toggle(self):
        """Show/hide the correct provider section and update self.provider."""
        selected = getattr(self, "_provider_var", None)
        if selected is None:
            return
        self.provider = selected.get()
        if self.provider == "ollama":
            if hasattr(self, "_openai_section"):
                self._openai_section.pack_forget()
            if hasattr(self, "_ollama_section"):
                self._ollama_section.pack(fill="x", pady=(8, 0))
            # Refresh model list immediately
            self._ollama_refresh_models()
        else:
            if hasattr(self, "_ollama_section"):
                self._ollama_section.pack_forget()
            if hasattr(self, "_openai_section"):
                self._openai_section.pack(fill="x")
        self._save_config()

    def _ollama_check_connection(self):
        """Ping Ollama server and update the status label."""
        url = self._ollama_url_var.get().strip()
        self.ollama_url = url
        if hasattr(self, "_ollama_status"):
            self._ollama_status.configure(text="Checking…", fg=C["warning"])
        self.root.update_idletasks()

        def _do():
            ok, err = JobApplicationAssistant.check_ollama_connection(url)
            def _apply():
                if ok:
                    models = JobApplicationAssistant.list_ollama_models(url)
                    n = len(models)
                    self._ollama_status.configure(
                        text=f"● Connected — {n} model{'s' if n != 1 else ''} installed",
                        fg=C["success"])
                    self._ollama_populate_list(models)
                else:
                    self._ollama_status.configure(text=f"✗ {err.splitlines()[0]}", fg=C["danger"])
            self.root.after(0, _apply)
        threading.Thread(target=_do, daemon=True).start()

    def _ollama_refresh_models(self):
        """Fetch installed models from Ollama and populate the listbox."""
        if not hasattr(self, "_ollama_listbox"):
            return
        url = self._ollama_url_var.get().strip()
        self.ollama_url = url

        def _do():
            ok, err = JobApplicationAssistant.check_ollama_connection(url)
            def _apply():
                if ok:
                    models = JobApplicationAssistant.list_ollama_models(url)
                    n = len(models)
                    if hasattr(self, "_ollama_status"):
                        self._ollama_status.configure(
                            text=f"● Connected — {n} model{'s' if n != 1 else ''} installed",
                            fg=C["success"])
                    self._ollama_populate_list(models)
                else:
                    if hasattr(self, "_ollama_status"):
                        self._ollama_status.configure(
                            text=f"✗ {err.splitlines()[0]}", fg=C["danger"])
                    self._ollama_listbox.delete(0, "end")
            self.root.after(0, _apply)
        threading.Thread(target=_do, daemon=True).start()

    def _ollama_populate_list(self, models: list[dict]):
        """Fill the Ollama model Listbox."""
        if not hasattr(self, "_ollama_listbox"):
            return
        self._ollama_listbox.delete(0, "end")
        for m in models:
            active = "● " if m["name"] == self.ollama_model else "  "
            self._ollama_listbox.insert(
                "end", f"{active}{m['name']:<30}  {m['size']:<8}  {m['modified']}")
        # Re-select the active model if still in list
        if self.ollama_model:
            for i, m in enumerate(models):
                if m["name"] == self.ollama_model:
                    self._ollama_listbox.selection_set(i)
                    break

    def _ollama_use_model(self):
        """Set the selected Ollama model as active and reinitialise."""
        if not hasattr(self, "_ollama_listbox"):
            return
        sel = self._ollama_listbox.curselection()
        if not sel:
            messagebox.showwarning("No Selection", "Select a model from the list first.")
            return
        raw = self._ollama_listbox.get(sel[0])
        # Strip the leading "● " or "  " marker and trailing size/date info
        model_name = raw.strip().lstrip("●").strip().split()[0]
        self.ollama_model = model_name
        self.model_var.set(model_name)
        self._save_config()
        if hasattr(self, "_ollama_model_lbl"):
            self._ollama_model_lbl.configure(
                text=f"Active model: {model_name}", fg=C["success"])
        # Refresh list to update the ● marker
        self._ollama_refresh_models()
        self._initialize_assistant()

    def _ollama_pull_model(self):
        """Open a dialog to pull a new Ollama model, showing streaming progress."""
        from tkinter import simpledialog
        name = simpledialog.askstring(
            "Pull Ollama Model",
            "Enter model name to pull:\n(e.g.  qwen2.5:7b  /  llama3.2  /  mistral-nemo)\n\n"
            "Browse all models at: ollama.com/library",
            parent=self.root)
        if not name:
            return
        name = name.strip()

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Pulling {name}")
        dlg.configure(bg=C["surface"])
        dlg.geometry("520x360")
        dlg.grab_set()

        tk.Label(dlg, text=f"Pulling model: {name}",
                 font=(FF, 10, "bold"), bg=C["surface"], fg=C["text"]).pack(
            anchor="w", padx=20, pady=(16, 4))
        tk.Label(dlg, text="This may take several minutes depending on model size.",
                 font=(FF, 8), bg=C["surface"], fg=C["text_muted"]).pack(anchor="w", padx=20)

        border = tk.Frame(dlg, bg=C["border_bright"], padx=1, pady=1)
        border.pack(fill="both", expand=True, padx=20, pady=10)
        log = tk.Text(border, font=(FM, 8), bg=C["input_bg"], fg=C["text"],
                      relief="flat", bd=0, wrap="word", padx=8, pady=6, state="disabled")
        log.pack(fill="both", expand=True)

        done_lbl = tk.Label(dlg, text="", font=(FF, 9, "bold"),
                            bg=C["surface"], fg=C["text_muted"])
        done_lbl.pack(pady=(0, 6))
        close_btn = self._btn(dlg, "Close", dlg.destroy, style="ghost")
        close_btn.pack(pady=(0, 12))
        close_btn.configure(state="disabled")

        def _append(line: str):
            log.configure(state="normal")
            log.insert("end", line + "\n")
            log.see("end")
            log.configure(state="disabled")

        def _pull():
            try:
                import requests as _req
                url = self._ollama_url_var.get().strip()
                resp = _req.post(f"{url}/api/pull",
                                 json={"name": name, "stream": True},
                                 stream=True, timeout=600)
                resp.raise_for_status()
                last_status = ""
                for raw_line in resp.iter_lines():
                    if raw_line:
                        try:
                            data = json.loads(raw_line)
                            status = data.get("status", "")
                            if status != last_status:
                                last_status = status
                                completed = data.get("completed", 0)
                                total     = data.get("total", 0)
                                if total and completed:
                                    pct = completed / total * 100
                                    line = f"{status}  ({pct:.0f}%)"
                                else:
                                    line = status
                                self.root.after(0, lambda l=line: _append(l))
                        except Exception:
                            pass
                def _done():
                    done_lbl.configure(text=f"✓ {name} pulled successfully", fg=C["success"])
                    close_btn.configure(state="normal")
                    self._ollama_refresh_models()
                self.root.after(0, _done)
            except Exception as e:
                def _err():
                    _append(f"ERROR: {e}")
                    done_lbl.configure(text="Pull failed — see log above", fg=C["danger"])
                    close_btn.configure(state="normal")
                self.root.after(0, _err)

        threading.Thread(target=_pull, daemon=True).start()

    def _select_model(self, model):
        self.model_var.set(model)
        for m, b in self._model_btns.items():
            b.configure(bg=C["accent"]    if m == model else C["surface3"],
                        fg=C["text_inv"]  if m == model else C["text_dim"])
        self._save_config()

    def _preview_cv_text(self):
        """Show the raw text extracted from the active CV in a scrollable dialog."""
        if not self.assistant:
            messagebox.showwarning("Not Initialized",
                                   "Initialize the assistant first to extract and preview CV text.")
            return
        text = self.assistant.cv_text.strip()
        chars = len(text)
        dlg = tk.Toplevel(self.root)
        dlg.title("CV Text Preview")
        dlg.configure(bg=C["surface"])
        dlg.geometry("720x560")
        dlg.grab_set()
        hdr = f"Extracted {chars:,} characters from {Path(self.cv_path.get()).name}"
        if chars < 500:
            hdr += "  ⚠  Very short — check PDF format"
        tk.Label(dlg, text=hdr, font=(FF, 9), bg=C["surface"],
                 fg=C["warning"] if chars < 500 else C["text_muted"],
                 anchor="w").pack(fill="x", padx=16, pady=(12, 6))
        border = tk.Frame(dlg, bg=C["border_bright"], padx=1, pady=1)
        border.pack(fill="both", expand=True, padx=16, pady=(0, 8))
        inner = tk.Frame(border, bg=C["input_bg"])
        inner.pack(fill="both", expand=True)
        t, sb = self._textarea(inner, height=28, mono=True)
        t.pack(side="left", fill="both", expand=True)
        sb.pack(side="left", fill="y")
        t.insert(1.0, text if text else "[No text extracted]")
        t.configure(state="disabled")
        self._btn(dlg, "Close", dlg.destroy, style="ghost").pack(pady=(0, 12))

    def _browse_cv(self):
        f = filedialog.askopenfilename(title="Select CV PDF",
                                       filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if f:
            self.cv_path.set(f)

    def _log(self, msg):
        def _do():
            self._setup_log.configure(state="normal")
            self._setup_log.insert("end", msg + "\n")
            self._setup_log.see("end")
            self._setup_log.configure(state="disabled")
        self.root.after(0, _do)

    def _auto_init(self):
        load_dotenv()
        cv_ready = Path(self.cv_path.get()).exists()
        if self.provider == "ollama":
            if cv_ready:
                self._initialize_assistant(silent=True)
        elif cv_ready and os.getenv("OPENAI_API_KEY"):
            self._initialize_assistant(silent=True)
        # Apply saved theme
        if getattr(self, "_saved_theme", "dark") == "light" and C["bg"] == C_DARK["bg"]:
            self._toggle_theme()
        # Apply saved font size if it differs from the widget default
        if self._font_size != 10:
            self._adjust_font_size(0)   # delta=0 triggers apply without changing the value
        self._restore_drafts()
        self._start_autosave()

    def _start_autosave(self):
        """Schedule periodic draft saves every 60 seconds."""
        self._autosave_drafts()
        self.root.after(60_000, self._start_autosave)

    def _autosave_drafts(self):
        drafts = {}
        jd_map = {
            "Evaluate":  "eval_job_text",
            "Generate":  "gen_job_text",
            "Q & A":     "qa_job_text",
            "Package":   "pkg_job_text",
            "Interview": "interview_job_text",
        }
        for page, attr in jd_map.items():
            w = getattr(self, attr, None)
            if w:
                try:
                    drafts[page] = w.get(1.0, "end").strip()
                except Exception:
                    pass
        try:
            with open("drafts.json", "w", encoding="utf-8") as f:
                json.dump({"saved": datetime.now().isoformat(), "drafts": drafts}, f, indent=2)
        except Exception:
            pass

    def _restore_drafts(self):
        if not Path("drafts.json").exists():
            return
        try:
            with open("drafts.json", "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return
        drafts = data.get("drafts", {})
        if not any(drafts.values()):
            return
        saved_at = data.get("saved", "")[:16].replace("T", " ")
        if not messagebox.askyesno(
                "Restore Drafts?",
                f"Unsaved JD drafts found from {saved_at}.\n\nRestore them?"):
            return
        jd_map = {
            "Evaluate":  "eval_job_text",
            "Generate":  "gen_job_text",
            "Q & A":     "qa_job_text",
            "Package":   "pkg_job_text",
            "Interview": "interview_job_text",
        }
        for page, attr in jd_map.items():
            text = drafts.get(page, "")
            w = getattr(self, attr, None)
            if w and text:
                try:
                    w.delete(1.0, "end")
                    w.insert(1.0, text)
                except Exception:
                    pass
        self._set_status("Drafts restored")

    def _initialize_assistant(self, silent=False):
        cv = self.cv_path.get()
        if not Path(cv).exists():
            messagebox.showerror("CV Not Found", f"File not found:\n{cv}")
            return

        self._setup_log.configure(state="normal")
        self._setup_log.delete(1.0, "end")
        self._setup_log.configure(state="disabled")
        self._log("Initializing …")
        self._set_status("Initializing assistant …")

        # ── Ollama path — no API key required ────────────────────────────
        if self.provider == "ollama":
            self.ollama_url = self._ollama_url_var.get().strip()
            if not self.ollama_model:
                self._log("ERROR: No Ollama model selected.")
                self._log("  → Select a model from the list, or pull one first.")
                if not silent:
                    messagebox.showerror("No Model Selected",
                                         "Select an Ollama model from the list first.\n"
                                         "If no models are installed, use 'Pull…' to download one.")
                return

            ok, err = JobApplicationAssistant.check_ollama_connection(self.ollama_url)
            if not ok:
                self._log(f"ERROR: {err}")
                if not silent:
                    messagebox.showerror("Ollama Not Running", err)
                return

            self._log(f"Provider:   Ollama  ({self.ollama_url})")
            self._log(f"Model:      {self.ollama_model}")

            def _do_ollama():
                try:
                    self.assistant = JobApplicationAssistant(
                        cv_path=cv,
                        provider="ollama",
                        base_url=f"{self.ollama_url}/v1",
                    )
                    cv_chars = len(self.assistant.cv_text.strip())
                    self._log(f"CV loaded:  {Path(cv).name}  ({cv_chars:,} chars extracted)")
                    if cv_chars < 500:
                        self._log("  WARNING: Very little text extracted from CV.")
                        self._log("  → Click 'Preview CV Text' to inspect.")
                    self._log_profile_source()
                    self._log("─" * 40)
                    self._log("Ready  (Ollama — running locally, no API cost)")
                    cv_label = self.active_cv_name or Path(cv).name
                    model_label = self.ollama_model
                    self._set_status(f"Ollama ready — {model_label}")
                    self.root.after(0, lambda: self._init_dot.configure(
                        text=f"● Ollama  ·  {model_label}", fg=C["success"]))
                    if hasattr(self, "_ollama_model_lbl"):
                        self.root.after(0, lambda: self._ollama_model_lbl.configure(
                            text=f"Active model: {model_label}", fg=C["success"]))
                    if not silent:
                        self.root.after(0, lambda: messagebox.showinfo(
                            "Ready", f"Ollama assistant ready!\nModel: {model_label}"))
                except Exception as e:
                    self._log(f"ERROR: {e}")
                    self._set_status("Initialization failed")
                    if not silent:
                        self.root.after(0, lambda: messagebox.showerror("Error", str(e)))

            threading.Thread(target=_do_ollama, daemon=True).start()
            return

        # ── OpenAI path ───────────────────────────────────────────────────
        load_dotenv()
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            self._log("ERROR: OPENAI_API_KEY not found in .env")
            if hasattr(self, "_api_status"):
                self._api_status.configure(text="Not configured", fg=C["danger"])
            if not silent:
                messagebox.showerror("API Key Missing",
                                     "Create a .env file with:\n\nOPENAI_API_KEY=your-key")
            return
        if hasattr(self, "_api_status"):
            self._api_status.configure(text="Found ●", fg=C["success"])
        self._log("API key found.")

        def _do():
            try:
                self.assistant = JobApplicationAssistant(cv_path=cv)
                cv_chars = len(self.assistant.cv_text.strip())
                self._log(f"CV loaded:  {Path(cv).name}  ({cv_chars:,} chars extracted)")
                if cv_chars < 500:
                    self._log("  WARNING: Very little text extracted from CV.")
                    self._log("  → Scanned PDF or unusual encoding?")
                    self._log("  → Click 'Preview CV Text' to inspect extraction.")
                self._log(f"Model:      {self.model_var.get()}")
                self._log_profile_source()
                self._log("─" * 40)
                self._log("Ready.")
                cv_label = self.active_cv_name or Path(cv).name
                self._set_status(f"Assistant ready — {cv_label}")
                self.root.after(0, lambda: self._init_dot.configure(
                    text=f"● Ready  ·  {cv_label}", fg=C["success"]))
                if hasattr(self, "_cv_active_lbl"):
                    self.root.after(0, lambda: self._cv_active_lbl.configure(
                        text=f"Active: {cv_label}  ({cv})", fg=C["success"]))
                if not silent:
                    self.root.after(0, lambda: messagebox.showinfo(
                        "Ready", "Assistant initialized successfully!"))
            except Exception as e:
                self._log(f"ERROR: {e}")
                self._set_status("Initialization failed")
                if not silent:
                    self.root.after(0, lambda: messagebox.showerror("Error", str(e)))

        threading.Thread(target=_do, daemon=True).start()

    def _log_profile_source(self):
        """Log which profile files are being used."""
        base = Path(__file__).parent
        _instr    = (base / PROFILE_INSTRUCTIONS_FILE).exists()
        _personal = (base / PROFILE_PERSONAL_FILE).exists()
        _legacy   = (base / PROFILE_FILE).exists()
        if _instr and _personal:
            label = "profile_instructions.md + profile_personal.md"
        elif _personal:
            label = "profile_personal.md"
        elif _legacy:
            label = "profile.md (legacy)"
        else:
            label = "built-in"
        self._log(f"Profile:    {label}")

    def _test_configuration(self):
        if not self._check_init():
            return
        self._log("\nRunning configuration test …")
        self._set_status("Testing …")

        def _do():
            try:
                r = self.assistant.evaluate_job_fit("ML Engineer, Python, PyTorch",
                                                     model=self.model_var.get())
                toks = r.get("tokens_used", 0)
                self._log(f"Test passed!  Tokens: {toks:,}")
                self._set_status(f"Test passed — {toks:,} tokens")
                self.root.after(0, lambda: messagebox.showinfo(
                    "Test Passed", f"Working!\nTokens used: {toks:,}"))
            except Exception as e:
                self._log(f"Test FAILED: {e}")
                self.root.after(0, lambda: messagebox.showerror("Test Failed", str(e)))

        threading.Thread(target=_do, daemon=True).start()

    # ═════════════════════════════════════════════════════════════════════════
    # Feature actions  (identical logic to v2)
    # ═════════════════════════════════════════════════════════════════════════

    def _evaluate_job_fit(self):
        if not self._check_init():
            return
        jd = self.eval_job_text.get(1.0, "end").strip()
        if not self._validate_jd(jd):
            return
        cb = self._stream_into(self.eval_results)
        self._set_status("Evaluating job fit …")

        def _do():
            try:
                r    = self.assistant.evaluate_job_fit(jd, model=self.model_var.get(), stream_callback=cb)
                toks = r.get("tokens_used", 0)
                text = r.get("evaluation", "")
                self._add_tokens(toks)
                self._set_status(f"Evaluation complete — {toks:,} tokens")
                self._add_history("Evaluation", text, toks, self.company_name.get(), self.role_title.get())
                self.root.after(0, self._refresh_eval_labels)
            except Exception as e:
                self._write(self.eval_results, f"Error: {e}")
                self._set_status("Error during evaluation")

        threading.Thread(target=_do, daemon=True).start()

    def _generate_cv_summary(self):
        if not self._check_init():
            return
        jd = self.gen_job_text.get(1.0, "end").strip()
        if not self._validate_jd(jd):
            return
        self._gen_eval_lbl._update()
        cb = self._stream_into(self.gen_output, prefix="CV SUMMARY\n" + "─" * 60 + "\n\n")
        self._set_status("Generating CV summary …")

        def _do():
            try:
                r    = self.assistant.generate_cv_summary(
                    jd, fit_evaluation=self.assistant._last_evaluation,
                    model=self.model_var.get(), stream_callback=cb)
                toks = r.get("tokens_used", 0)
                self._add_tokens(toks)
                self._set_status(f"CV summary generated — {toks:,} tokens")
                self._add_history("CV Summary", r.get("summary", ""), toks,
                                  self.company_name.get(), self.role_title.get())
            except Exception as e:
                self._write(self.gen_output, f"Error: {e}")
                self._set_status("Error generating CV summary")

        threading.Thread(target=_do, daemon=True).start()

    def _generate_cover_letter(self, tone: str = "hybrid"):
        if not self._check_init():
            return
        jd = self.gen_job_text.get(1.0, "end").strip()
        if not self._validate_jd(jd):
            return
        company = self.company_name.get().strip() or None
        role    = self.role_title.get().strip()    or None
        self._gen_eval_lbl._update()
        tone_label = {"hybrid": "COVER LETTER (Hybrid)",
                      "research": "COVER LETTER (Research-led)",
                      "engineering": "COVER LETTER (Engineering-led)"}.get(tone, "COVER LETTER")
        cb = self._stream_into(self.gen_output, prefix=f"{tone_label}\n" + "─" * 60 + "\n\n")
        self._set_status(f"Generating cover letter [{tone}] …")

        def _do():
            try:
                r    = self.assistant.generate_cover_letter(
                    jd, fit_evaluation=self.assistant._last_evaluation,
                    company_name=company, role_title=role, tone=tone,
                    model=self.model_var.get(), stream_callback=cb)
                toks = r.get("tokens_used", 0)
                self._add_tokens(toks)
                self._set_status(f"Cover letter [{tone}] generated — {toks:,} tokens")
                self._add_history(f"Cover Letter ({tone.title()})",
                                  r.get("cover_letter", ""), toks, company, role)
            except Exception as e:
                self._write(self.gen_output, f"Error: {e}")
                self._set_status("Error generating cover letter")

        threading.Thread(target=_do, daemon=True).start()

    def _generate_linkedin_message(self):
        if not self._check_init():
            return
        jd = self.gen_job_text.get(1.0, "end").strip()
        if not self._validate_jd(jd):
            return
        recruiter = self.recruiter_name.get().strip() or None
        cb = self._stream_into(self.gen_output, prefix="LINKEDIN MESSAGE\n" + "─" * 60 + "\n\n")
        self._set_status("Generating LinkedIn message …")

        def _do():
            try:
                r    = self.assistant.generate_linkedin_message(
                    jd, recruiter_name=recruiter,
                    fit_evaluation=self.assistant._last_evaluation,
                    model=self.model_var.get(), stream_callback=cb)
                toks = r.get("tokens_used", 0)
                self._add_tokens(toks)
                self._set_status(f"LinkedIn message generated — {toks:,} tokens")
                self._add_history("LinkedIn", r.get("linkedin_message", ""), toks,
                                  self.company_name.get(), self.role_title.get())
            except Exception as e:
                self._write(self.gen_output, f"Error: {e}")
                self._set_status("Error generating LinkedIn message")

        threading.Thread(target=_do, daemon=True).start()

    def _open_salary_negotiation(self):
        """Open a dialog to paste an offer, then generate the negotiation brief."""
        if not self._check_init():
            return
        dlg = tk.Toplevel(self.root)
        dlg.title("Salary Negotiation Coach")
        dlg.configure(bg=C["surface"])
        dlg.geometry("580x420")
        dlg.resizable(False, True)
        dlg.grab_set()

        tk.Label(dlg, text="Paste the offer details below:",
                 font=(FF, 10, "bold"), bg=C["surface"], fg=C["text"]).pack(
            anchor="w", padx=20, pady=(16, 4))
        tk.Label(dlg,
                 text="Include: base salary, bonus, equity, benefits, start date — whatever was offered.",
                 font=(FF, 8), bg=C["surface"], fg=C["text_muted"], wraplength=540).pack(
            anchor="w", padx=20, pady=(0, 8))

        border = tk.Frame(dlg, bg=C["border_bright"], padx=1, pady=1)
        border.pack(fill="both", expand=True, padx=20)
        offer_text = tk.Text(border, font=(FF, 9), bg=C["input_bg"], fg=C["text"],
                             insertbackground=C["accent_soft"], relief="flat",
                             bd=0, wrap="word", padx=8, pady=6, undo=True)
        offer_text.pack(fill="both", expand=True)
        offer_text.bind("<FocusIn>",  lambda e: border.configure(bg=C["accent"]))
        offer_text.bind("<FocusOut>", lambda e: border.configure(bg=C["border_bright"]))

        btn_row = tk.Frame(dlg, bg=C["surface"])
        btn_row.pack(fill="x", padx=20, pady=(10, 16))

        def _run():
            offer = offer_text.get(1.0, "end").strip()
            if not offer:
                messagebox.showwarning("Empty", "Paste the offer details first.", parent=dlg)
                return
            dlg.destroy()
            jd = self.gen_job_text.get(1.0, "end").strip() or None
            cb = self._stream_into(self.gen_output,
                                   prefix="SALARY NEGOTIATION BRIEF\n" + "─" * 60 + "\n\n")
            self._set_status("Generating negotiation brief …")

            def _do():
                try:
                    r    = self.assistant.generate_salary_negotiation(
                        offer, job_description=jd,
                        model=self.model_var.get(), stream_callback=cb)
                    toks = r.get("tokens_used", 0)
                    self._add_tokens(toks)
                    self._set_status(f"Negotiation brief ready — {toks:,} tokens")
                    self._add_history("Salary Negotiation", r.get("salary_negotiation", ""),
                                      toks, self.company_name.get(), self.role_title.get())
                except Exception as e:
                    self._write(self.gen_output, f"Error: {e}")
                    self._set_status("Error generating negotiation brief")

            threading.Thread(target=_do, daemon=True).start()

        self._btn(btn_row, "Generate Negotiation Brief", _run,    style="primary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Cancel",                     dlg.destroy, style="ghost").pack(side="left")

    def _run_ats_check(self):
        if not self._check_init():
            return
        jd = self.gen_job_text.get(1.0, "end").strip()
        if not jd:
            messagebox.showwarning("Input Required", "Please paste a job description first.")
            return
        content = self.gen_output.get(1.0, "end").strip()
        if not content:
            messagebox.showwarning("No Content",
                                   "Generate a CV summary or cover letter first,\n"
                                   "then click ATS Check to analyse it.")
            return
        cb = self._stream_into(self.gen_output, prefix="ATS KEYWORD ANALYSIS\n" + "─" * 60 + "\n\n")
        self._set_status("Running ATS analysis …")

        def _do():
            try:
                r    = self.assistant.analyze_ats_fit(
                    jd, content, model=self.model_var.get(), stream_callback=cb)
                toks = r.get("tokens_used", 0)
                self._add_tokens(toks)
                self._set_status(f"ATS analysis complete — {toks:,} tokens")
                self._add_history("ATS Check", r.get("ats_analysis", ""), toks,
                                  self.company_name.get(), self.role_title.get())
            except Exception as e:
                self._write(self.gen_output, f"Error: {e}")
                self._set_status("Error running ATS analysis")

        threading.Thread(target=_do, daemon=True).start()

    def _answer_question(self):
        if not self._check_init():
            return
        jd       = self.qa_job_text.get(1.0, "end").strip()
        question = self.qa_question.get(1.0, "end").strip()
        if not self._validate_jd(jd):
            return
        if not question:
            messagebox.showwarning("Input Required", "Please enter your question.")
            return
        cb = self._stream_into(self.qa_output)
        self._set_status("Answering question …")

        def _do():
            try:
                r    = self.assistant.answer_application_question(
                    jd, question, fit_evaluation=self.assistant._last_evaluation,
                    model=self.model_var.get(), stream_callback=cb)
                toks = r.get("tokens_used", 0)
                self._add_tokens(toks)
                self._set_status(f"Answer generated — {toks:,} tokens")
                display = f"QUESTION\n{'─'*60}\n{question}\n\nANSWER\n{'─'*60}\n\n{r.get('answer','')}"
                self._add_history("Q&A", display, toks, self.company_name.get(), self.role_title.get())
            except Exception as e:
                self._write(self.qa_output, f"Error: {e}")
                self._set_status("Error during Q&A")

        threading.Thread(target=_do, daemon=True).start()

    def _generate_interview_prep(self):
        if not self._check_init():
            return
        jd = self.interview_job_text.get(1.0, "end").strip()
        if not self._validate_jd(jd):
            return
        self._int_eval_lbl._update()
        cb = self._stream_into(self.interview_output)
        self._set_status("Generating interview prep …")

        def _do():
            try:
                r    = self.assistant.generate_interview_prep(
                    jd, fit_evaluation=self.assistant._last_evaluation,
                    model=self.model_var.get(), stream_callback=cb)
                toks = r.get("tokens_used", 0)
                self._add_tokens(toks)
                self._set_status(f"Interview prep generated — {toks:,} tokens")
                self._add_history("Interview Prep", r.get("interview_prep", ""), toks,
                                  self.company_name.get(), self.role_title.get())
            except Exception as e:
                self._write(self.interview_output, f"Error: {e}")
                self._set_status("Error generating interview prep")

        threading.Thread(target=_do, daemon=True).start()

    def _open_quiz_mode(self):
        """Parse interview prep output into Q&A flashcards and run a quiz."""
        raw = self.interview_output.get(1.0, "end").strip()
        if not raw:
            messagebox.showwarning("No Content",
                                   "Generate interview prep first, then launch Quiz Mode.")
            return

        # Extract question/answer pairs — any line ending with "?" is a question;
        # everything up to the next question or section header is the answer.
        lines  = raw.splitlines()
        pairs  = []
        q, a_lines = None, []
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            if stripped.endswith("?") and len(stripped) > 10:
                if q:
                    pairs.append((q, "\n".join(a_lines).strip()))
                q, a_lines = stripped, []
            elif q:
                a_lines.append(line)
        if q:
            pairs.append((q, "\n".join(a_lines).strip()))

        if not pairs:
            messagebox.showwarning("No Questions Found",
                                   "Couldn't parse questions from the interview prep.\n"
                                   "Questions must end with '?'.")
            return

        # ── Quiz window ───────────────────────────────────────────────────────
        dlg = tk.Toplevel(self.root)
        dlg.title(f"Quiz Mode — {len(pairs)} questions")
        dlg.configure(bg=C["bg"])
        dlg.geometry("700x520")
        dlg.grab_set()

        state = {"idx": 0, "revealed": False, "score": 0}

        # Header
        hdr = tk.Frame(dlg, bg=C["surface"], pady=10)
        hdr.pack(fill="x")
        progress_lbl = tk.Label(hdr, text="", font=(FF, 9), bg=C["surface"], fg=C["text_muted"])
        progress_lbl.pack(side="left", padx=16)
        score_lbl = tk.Label(hdr, text="Score: 0 / 0", font=(FF, 9, "bold"),
                             bg=C["surface"], fg=C["accent_soft"])
        score_lbl.pack(side="right", padx=16)

        # Question
        q_frame = tk.Frame(dlg, bg=C["surface2"], padx=20, pady=16)
        q_frame.pack(fill="x", padx=20, pady=(16, 8))
        q_lbl = tk.Label(q_frame, text="", font=(FF, 11, "bold"), bg=C["surface2"],
                         fg=C["text"], wraplength=620, justify="left")
        q_lbl.pack(anchor="w")

        # Your answer
        tk.Label(dlg, text="Your answer:", font=(FF, 8), bg=C["bg"],
                 fg=C["text_muted"]).pack(anchor="w", padx=20, pady=(0, 4))
        ans_border = tk.Frame(dlg, bg=C["border_bright"], padx=1, pady=1)
        ans_border.pack(fill="x", padx=20)
        user_ans = tk.Text(ans_border, height=4, font=(FF, 9), bg=C["input_bg"], fg=C["text"],
                           insertbackground=C["accent_soft"], relief="flat", bd=0,
                           wrap="word", padx=8, pady=6, undo=True)
        user_ans.pack(fill="x")

        # Model answer (hidden until revealed)
        reveal_lbl = tk.Label(dlg, text="", font=(FF, 8), bg=C["bg"], fg=C["text_muted"])
        reveal_lbl.pack(anchor="w", padx=20, pady=(8, 2))
        ans_text_border = tk.Frame(dlg, bg=C["border_bright"], padx=1, pady=1)
        ans_text_border.pack(fill="x", padx=20)
        model_ans = tk.Text(ans_text_border, height=5, font=(FF, 9), bg=C["input_bg"],
                            fg=C["text_dim"], relief="flat", bd=0, wrap="word",
                            padx=8, pady=6, state="disabled")
        model_ans.pack(fill="x")
        ans_text_border.pack_forget()   # hidden until revealed

        btn_row = tk.Frame(dlg, bg=C["bg"])
        btn_row.pack(pady=(10, 0), padx=20, fill="x")

        def load_question():
            idx = state["idx"]
            state["revealed"] = False
            q, _ = pairs[idx]
            progress_lbl.configure(text=f"Question {idx + 1} of {len(pairs)}")
            q_lbl.configure(text=q)
            user_ans.delete(1.0, "end")
            model_ans.configure(state="normal")
            model_ans.delete(1.0, "end")
            model_ans.configure(state="disabled")
            reveal_lbl.configure(text="")
            ans_text_border.pack_forget()
            reveal_btn.configure(state="normal")
            next_btn.configure(state="disabled")
            good_btn.configure(state="disabled")

        def reveal():
            if state["revealed"]:
                return
            state["revealed"] = True
            _, answer = pairs[state["idx"]]
            model_ans.configure(state="normal")
            model_ans.delete(1.0, "end")
            model_ans.insert(1.0, answer if answer else "[No model answer found]")
            model_ans.configure(state="disabled")
            reveal_lbl.configure(text="Model answer:")
            ans_text_border.pack(fill="x", padx=20)
            reveal_btn.configure(state="disabled")
            next_btn.configure(state="normal")
            good_btn.configure(state="normal")

        def mark_good():
            state["score"] += 1
            advance()

        def advance():
            total_done = state["idx"] + 1
            score_lbl.configure(text=f"Score: {state['score']} / {total_done}")
            if state["idx"] < len(pairs) - 1:
                state["idx"] += 1
                load_question()
            else:
                pct = state["score"] / len(pairs) * 100
                messagebox.showinfo("Quiz Complete",
                                    f"Finished!\n\nScore: {state['score']} / {len(pairs)} "
                                    f"({pct:.0f}%)",
                                    parent=dlg)
                dlg.destroy()

        reveal_btn = self._btn(btn_row, "Reveal Answer", reveal,     style="primary")
        reveal_btn.pack(side="left", padx=(0, 8))
        good_btn   = self._btn(btn_row, "✓ Got it",     mark_good,   style="success")
        good_btn.pack(side="left", padx=(0, 8))
        good_btn.configure(state="disabled")
        next_btn   = self._btn(btn_row, "→ Skip / Next", advance,    style="ghost")
        next_btn.pack(side="left")
        next_btn.configure(state="disabled")
        self._btn(btn_row, "Exit Quiz", dlg.destroy, style="danger").pack(side="right")

        load_question()

    def _generate_followup_email(self):
        if not self._check_init():
            return
        jd = self.interview_job_text.get(1.0, "end").strip()
        if not jd:
            messagebox.showwarning("Input Required",
                                   "Paste a job description on the left before generating a follow-up email.")
            return
        interviewer = self._followup_name_var.get().strip() or None
        notes       = self.followup_notes_text.get(1.0, "end").strip() or None
        cb = self._stream_into(self.interview_output)
        self._set_status("Generating follow-up email …")

        def _do():
            try:
                r    = self.assistant.generate_followup_email(
                    jd, interviewer_name=interviewer, interview_notes=notes,
                    model=self.model_var.get(), stream_callback=cb)
                toks = r.get("tokens_used", 0)
                self._add_tokens(toks)
                self._set_status(f"Follow-up email generated — {toks:,} tokens")
                self._add_history("Follow-Up Email", r.get("followup_email", ""), toks,
                                  self.company_name.get(), self.role_title.get())
            except Exception as e:
                self._write(self.interview_output, f"Error: {e}")
                self._set_status("Error generating follow-up email")

        threading.Thread(target=_do, daemon=True).start()

    def _generate_package(self):
        if not self._check_init():
            return
        jd = self.pkg_job_text.get(1.0, "end").strip()
        if not self._validate_jd(jd):
            return
        company = self.company_name.get().strip() or None
        role    = self.role_title.get().strip()    or None
        self.pkg_output.configure(state="normal")
        self.pkg_output.delete(1.0, "end")
        self.pkg_output.configure(state="disabled")
        self._set_status("Generating package — step 1/3: Evaluation …")
        self._pkg_running = True

        def _animate():
            import time
            frames = ["   ", ".  ", ".. ", "..."]
            i = 0
            while self._pkg_running:
                dot = frames[i % 4]
                self.root.after(0, lambda d=dot: self._progress_lbl.configure(text="Working" + d))
                i += 1
                time.sleep(0.5)
            self.root.after(0, lambda: self._progress_lbl.configure(text=""))

        threading.Thread(target=_animate, daemon=True).start()

        def _do():
            total_tokens = 0
            eval_text = ""
            try:
                self._append_to(self.pkg_output,
                                f"COMPLETE APPLICATION PACKAGE\n{'═'*60}\n\n{'═'*60}\n"
                                f"JOB FIT EVALUATION\n{'─'*60}\n\n")
                cb1 = self._make_append_callback(self.pkg_output)
                r1  = self.assistant.evaluate_job_fit(jd, model=self.model_var.get(), stream_callback=cb1)
                eval_text     = r1.get("evaluation", "")
                total_tokens += r1.get("tokens_used", 0)
                self._set_status("Generating package — step 2/3: CV Summary …")

                self._append_to(self.pkg_output,
                                f"\n\n{'═'*60}\nCV SUMMARY\n{'─'*60}\n\n")
                cb2 = self._make_append_callback(self.pkg_output)
                r2  = self.assistant.generate_cv_summary(
                    jd, fit_evaluation=eval_text,
                    model=self.model_var.get(), stream_callback=cb2)
                total_tokens += r2.get("tokens_used", 0)
                self._set_status("Generating package — step 3/3: Cover Letter …")

                self._append_to(self.pkg_output,
                                f"\n\n{'═'*60}\nCOVER LETTER\n{'─'*60}\n\n")
                cb3 = self._make_append_callback(self.pkg_output)
                r3  = self.assistant.generate_cover_letter(
                    jd, fit_evaluation=eval_text,
                    company_name=company, role_title=role,
                    model=self.model_var.get(), stream_callback=cb3)
                total_tokens += r3.get("tokens_used", 0)

                self._pkg_running = False
                self._add_tokens(total_tokens)
                self._set_status(f"Package complete — {total_tokens:,} tokens")

                def _finalise():
                    full_text = self.pkg_output.get(1.0, "end").strip()
                    self._add_history("Package", full_text, total_tokens, company, role)
                    messagebox.showinfo("Done", f"Package generated!\n\nTotal tokens: {total_tokens:,}")

                self.root.after(0, _finalise)
            except Exception as e:
                self._pkg_running = False
                self._write(self.pkg_output, f"Error: {e}")
                self._set_status("Error during package generation")

        threading.Thread(target=_do, daemon=True).start()

    def _auto_extract_job_details(self):
        if not self._check_init():
            return
        jd_widget = self._jd_widgets.get(self._current_page)
        if jd_widget is None:
            messagebox.showwarning("No JD", "Navigate to a page with a job description first.")
            return
        jd = jd_widget.get(1.0, "end").strip()
        if not jd:
            messagebox.showwarning("No JD", "Paste a job description first.")
            return
        self._set_status("Extracting job details …")

        def _do():
            try:
                r = self.assistant.extract_job_details(jd, model=self.model_var.get())
                def _apply():
                    if r.get("company"): self.company_name.set(r["company"])
                    if r.get("role"):    self.role_title.set(r["role"])
                    self._set_status(f"Extracted: {r.get('company','?')} — {r.get('role','?')}")
                self.root.after(0, _apply)
            except Exception as e:
                self._set_status(f"Extraction error: {e}")

        threading.Thread(target=_do, daemon=True).start()

    # ─────────────────────────────────────────────────────────────────────────
    # Profile editor
    # ─────────────────────────────────────────────────────────────────────────

    def _open_profile_wizard(self):
        """Open the guided profile setup wizard."""
        def _on_save():
            self._load_profile_editor()
            if self.assistant:
                self._initialize_assistant(silent=True)
            self._set_status("Profile wizard complete — personal profile saved")
        ProfileWizard(self.root, on_save=_on_save)

    def _load_profile_editor(self):
        """Reload both profile editors from disk."""
        base = Path(__file__).parent

        # Personal profile editor
        if hasattr(self, "profile_personal_editor"):
            personal_path = base / PROFILE_PERSONAL_FILE
            try:
                personal_content = personal_path.read_text(encoding="utf-8") if personal_path.exists() else ""
            except Exception as e:
                personal_content = f"# Error reading {PROFILE_PERSONAL_FILE}\n# {e}"
            self.profile_personal_editor.delete(1.0, "end")
            self.profile_personal_editor.insert(1.0, personal_content)

        # Instructions editor
        if hasattr(self, "profile_instr_editor"):
            instr_path = base / PROFILE_INSTRUCTIONS_FILE
            try:
                instr_content = instr_path.read_text(encoding="utf-8") if instr_path.exists() else ""
            except Exception as e:
                instr_content = f"# Error reading {PROFILE_INSTRUCTIONS_FILE}\n# {e}"
            self.profile_instr_editor.delete(1.0, "end")
            self.profile_instr_editor.insert(1.0, instr_content)

        if hasattr(self, "_profile_status"):
            self._profile_status.configure(
                text=f"Loaded  {datetime.now().strftime('%H:%M:%S')}",
                fg=C["text_muted"])

    def _save_profile(self):
        """Save both profile files and reload the assistant."""
        base = Path(__file__).parent
        personal_path = base / PROFILE_PERSONAL_FILE
        instr_path    = base / PROFILE_INSTRUCTIONS_FILE

        personal_content = (self.profile_personal_editor.get(1.0, "end-1c")
                            if hasattr(self, "profile_personal_editor") else "")
        instr_content    = (self.profile_instr_editor.get(1.0, "end-1c")
                            if hasattr(self, "profile_instr_editor") else "")

        if not personal_content.strip() and not instr_content.strip():
            messagebox.showwarning("Empty Profile", "Both editors are empty. Nothing was saved.")
            return

        errors = []
        if personal_content.strip():
            try:
                personal_path.write_text(personal_content, encoding="utf-8")
            except Exception as e:
                errors.append(f"{PROFILE_PERSONAL_FILE}: {e}")

        if instr_content.strip():
            try:
                instr_path.write_text(instr_content, encoding="utf-8")
            except Exception as e:
                errors.append(f"{PROFILE_INSTRUCTIONS_FILE}: {e}")

        if errors:
            messagebox.showerror("Save Error", "Could not save:\n" + "\n".join(errors))
            return

        self._profile_status.configure(
            text=f"Saved  {datetime.now().strftime('%H:%M:%S')}",
            fg=C["success"])
        self._set_status("Profile saved")
        if self.assistant:
            self._initialize_assistant(silent=True)

        # Placeholder completeness check — personal tab only
        placeholders = re.findall(r'\[YOUR[^\]]*\]', personal_content)
        if placeholders:
            unique = list(dict.fromkeys(placeholders))  # deduplicate, preserve order
            preview = ", ".join(unique[:4])
            if len(unique) > 4:
                preview += f"  ... +{len(unique) - 4} more"
            self._profile_warn.configure(
                text=f"  {len(unique)} placeholder(s) still unfilled: {preview}")
            self._profile_warn.pack(fill="x", pady=(0, 6),
                                    after=self._profile_info_outer)
        else:
            self._profile_warn.pack_forget()

    # ═════════════════════════════════════════════════════════════════════════
    # Application Tracker
    # ═════════════════════════════════════════════════════════════════════════

    # ─────────────────────────────────────────────────────────────────────────
    # Config (multi-CV, preferences)
    # ─────────────────────────────────────────────────────────────────────────

    def _load_config(self):
        """Load config.json; restore CV profiles, model, and window geometry."""
        path = Path(CONFIG_FILE)
        if path.exists():
            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                self.cv_profiles    = cfg.get("cv_profiles", [])
                self.active_cv_name = cfg.get("active_cv_name", "")
                # Restore active CV path
                active = next((p for p in self.cv_profiles
                               if p["name"] == self.active_cv_name), None)
                if active:
                    self.cv_path.set(active["path"])
                # Restore model
                saved_model = cfg.get("model")
                if saved_model:
                    self.model_var.set(saved_model)
                # Restore window geometry
                geom = cfg.get("geometry")
                if geom:
                    try:
                        self.root.geometry(geom)
                    except Exception:
                        pass
                # Restore font size (applied after UI is built via _auto_init)
                self._font_size   = int(cfg.get("font_size", 10))
                self._saved_theme = cfg.get("theme", "dark")
                # Restore Ollama settings
                self.provider     = cfg.get("provider", "openai")
                self.ollama_url   = cfg.get("ollama_url", "http://localhost:11434")
                self.ollama_model = cfg.get("ollama_model", "")
                self._ollama_url_var.set(self.ollama_url)
                return
            except Exception:
                pass
        # No config yet — seed from env / default
        default_path = os.getenv("CV_PATH", "cv.pdf")
        default_name = Path(default_path).stem
        self.cv_profiles    = [{"name": default_name, "path": default_path}]
        self.active_cv_name = default_name

    def _save_config(self):
        try:
            cfg = {
                "cv_profiles":    self.cv_profiles,
                "active_cv_name": self.active_cv_name,
                "model":          self.model_var.get(),
                "geometry":       self.root.geometry(),
                "font_size":      self._font_size,
                "theme":          "light" if C["bg"] == C_LIGHT["bg"] else "dark",
                "provider":       self.provider,
                "ollama_url":     self.ollama_url,
                "ollama_model":   self.ollama_model,
            }
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(cfg, f, indent=2)
        except Exception as e:
            print(f"[Config] save error: {e}")

    def _cv_refresh_list(self):
        if not hasattr(self, "_cv_listbox"):
            return
        self._cv_listbox.delete(0, "end")
        for p in self.cv_profiles:
            marker = "● " if p["name"] == self.active_cv_name else "  "
            self._cv_listbox.insert("end", f"{marker}{p['name']}  —  {p['path']}")
        if hasattr(self, "_cv_active_lbl"):
            active = next((p for p in self.cv_profiles
                           if p["name"] == self.active_cv_name), None)
            if active:
                self._cv_active_lbl.configure(
                    text=f"Active: {active['name']}  ({active['path']})",
                    fg=C["text_muted"])

    def _cv_add(self):
        from tkinter import simpledialog
        path = filedialog.askopenfilename(
            title="Select CV PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if not path:
            return
        name = simpledialog.askstring(
            "CV Name",
            "Give this CV a short name (e.g. 'Research CV', 'Engineering CV'):",
            initialvalue=Path(path).stem,
            parent=self.root)
        if not name:
            return
        name = name.strip()
        # Update existing or add new
        for p in self.cv_profiles:
            if p["name"] == name:
                p["path"] = path
                self._save_config()
                self._cv_refresh_list()
                return
        self.cv_profiles.append({"name": name, "path": path})
        self._save_config()
        self._cv_refresh_list()

    def _cv_remove(self):
        sel = self._cv_listbox.curselection()
        if not sel:
            messagebox.showwarning("No Selection", "Select a CV to remove.")
            return
        idx  = sel[0]
        name = self.cv_profiles[idx]["name"]
        if messagebox.askyesno("Remove CV", f"Remove '{name}' from the list?\n(The file itself is not deleted.)"):
            self.cv_profiles.pop(idx)
            if self.active_cv_name == name and self.cv_profiles:
                self.active_cv_name = self.cv_profiles[0]["name"]
                self.cv_path.set(self.cv_profiles[0]["path"])
            self._save_config()
            self._cv_refresh_list()

    def _cv_use_selected(self):
        sel = self._cv_listbox.curselection()
        if not sel:
            messagebox.showwarning("No Selection", "Select a CV to load.")
            return
        profile = self.cv_profiles[sel[0]]
        if not Path(profile["path"]).exists():
            messagebox.showerror("File Not Found",
                                 f"CV file not found:\n{profile['path']}\n\nUpdate the path by re-adding this CV.")
            return
        self.active_cv_name = profile["name"]
        self.cv_path.set(profile["path"])
        self._save_config()
        self._cv_refresh_list()
        self._initialize_assistant()

    def _load_applications(self):
        if Path(APPLICATIONS_FILE).exists():
            try:
                with open(APPLICATIONS_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return []
        return []

    def _save_applications(self):
        try:
            with open(APPLICATIONS_FILE, "w", encoding="utf-8") as f:
                json.dump(self.applications, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"[Tracker] save error: {e}")

    def _configure_tracker_style(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Tracker.Treeview",
                        background=C["surface"], foreground=C["text"],
                        fieldbackground=C["surface"], borderwidth=0,
                        rowheight=30, font=(FF, 9))
        style.configure("Tracker.Treeview.Heading",
                        background=C["surface2"], foreground=C["accent_soft"],
                        font=(FF, 9, "bold"), borderwidth=0, relief="flat")
        style.map("Tracker.Treeview",
                  background=[("selected", C["surface3"])],
                  foreground=[("selected", C["accent_soft"])])

    def _refresh_tracker_table(self):
        if not hasattr(self, "_tracker_tree"):
            return
        # Count by status and update stats bar
        counts = {s: 0 for s in STATUSES}
        for app in self.applications:
            s = app.get("status", "Applied")
            if s in counts:
                counts[s] += 1
        for status, (cell, cnt_lbl, name_lbl, dot) in self._stats_count_labels.items():
            cnt_lbl.configure(text=str(counts.get(status, 0)))
            # Highlight active filter
            is_active = self._tracker_filter_status == status
            _, fg = STATUS_ROW_COLORS[status]
            bg = C["surface2"] if is_active else C["bg"]
            for w in (cell, cnt_lbl, name_lbl, dot):
                w.configure(bg=bg)

        AGE_THRESHOLD_DAYS = 14
        ACTIVE_STATUSES = {"Watching", "Applied", "Phone Screen", "Interview", "Final Round"}
        today = datetime.now().date()

        self._tracker_tree.delete(*self._tracker_tree.get_children())
        for app in self.applications:
            status = app.get("status", "Applied")
            if self._tracker_filter_status and status != self._tracker_filter_status:
                continue
            tag = status.replace(" ", "_")

            # Age indicator: flag active applications sitting >14 days
            date_str = app.get("date_applied", "")
            aged = False
            if status in ACTIVE_STATUSES and date_str:
                try:
                    applied = datetime.strptime(date_str, "%Y-%m-%d").date()
                    if (today - applied).days > AGE_THRESHOLD_DAYS:
                        aged = True
                except ValueError:
                    pass
            display_date = f"⚑ {date_str}" if aged else date_str

            # "● Status" gives a badge-like appearance via the coloured status text
            self._tracker_tree.insert("", "end", iid=app["id"],
                values=(
                    display_date,
                    app.get("company",      ""),
                    app.get("position",     ""),
                    app.get("location",     ""),
                    f"● {status}",
                    app.get("salary_jd",    ""),
                    app.get("salary_req",   ""),
                    app.get("source",       ""),
                ),
                tags=(tag,))

    def _tracker_filter_by_status(self, status):
        """Toggle table filter to a single status; click again to clear."""
        if self._tracker_filter_status == status:
            self._tracker_filter_status = None
        else:
            self._tracker_filter_status = status
        self._refresh_tracker_table()

    def _tracker_sort_by(self, col):
        if self._tracker_sort_col == col:
            self._tracker_sort_rev = not self._tracker_sort_rev
        else:
            self._tracker_sort_col = col
            self._tracker_sort_rev = False
        col_key = {"date": "date_applied", "company": "company", "position": "position",
                   "location": "location", "status": "status", "salary_jd": "salary_jd",
                   "salary_req": "salary_req", "source": "source"}.get(col, col)
        self.applications.sort(key=lambda a: a.get(col_key, "").lower(),
                               reverse=self._tracker_sort_rev)
        self._save_applications()
        self._refresh_tracker_table()
        for c_id, c_hdr, _ in TRACKER_COLS:
            arrow = (" ▼" if self._tracker_sort_rev else " ▲") if c_id == col else ""
            self._tracker_tree.heading(c_id, text=c_hdr + arrow)

    def _on_tracker_select(self, _event=None):
        sel = self._tracker_tree.selection()
        if not sel:
            return
        app_id = sel[0]
        self._tracker_selected_id = app_id
        app = next((a for a in self.applications if a["id"] == app_id), None)
        if not app:
            return
        self._trk_company.set(app.get("company",         ""))
        self._trk_position.set(app.get("position",       ""))
        self._trk_location.set(app.get("location",       ""))
        self._trk_date.set(app.get("date_applied",       ""))
        self._trk_date_interview.set(app.get("date_interview", ""))
        self._trk_date_decision.set(app.get("date_decision",  ""))
        self._trk_source.set(app.get("source",           ""))
        peak = app.get("peak_stage", "")
        self._trk_peak_stage.set(f"● {peak}" if peak else "")
        self._trk_status_var.set(app.get("status",       "Applied"))
        self._trk_salary_jd.set(app.get("salary_jd",    ""))
        self._trk_salary_req.set(app.get("salary_req",  ""))
        self._trk_notes.delete(1.0, "end")
        self._trk_notes.insert(1.0, app.get("notes", ""))
        if hasattr(self, "_trk_status_lbl"):
            self._trk_status_lbl.configure(
                text=f"Editing: {app.get('company','')} — {app.get('position','')}",
                fg=C["text_dim"])

    def _tracker_new_entry(self):
        self._tracker_selected_id = None
        self._trk_company.set("")
        self._trk_position.set("")
        self._trk_location.set("")
        self._trk_date.set(datetime.now().strftime("%Y-%m-%d"))
        self._trk_date_interview.set("")
        self._trk_date_decision.set("")
        self._trk_source.set("")
        self._trk_peak_stage.set("")
        self._trk_status_var.set("Applied")
        self._trk_salary_jd.set("")
        self._trk_salary_req.set("")
        if hasattr(self, "_trk_notes"):
            self._trk_notes.delete(1.0, "end")
        if hasattr(self, "_tracker_tree"):
            for item in self._tracker_tree.selection():
                self._tracker_tree.selection_remove(item)
        if hasattr(self, "_trk_status_lbl"):
            self._trk_status_lbl.configure(
                text="New entry — fill in details and click Save Entry",
                fg=C["text_muted"])

    def _tracker_save_entry(self):
        company  = self._trk_company.get().strip()
        position = self._trk_position.get().strip()
        if not company and not position:
            messagebox.showwarning("Empty Entry", "Enter at least a Company or Position.")
            return
        new_status = self._trk_status_var.get()

        # Determine peak_stage: advance only on forward progression; preserve on terminal outcomes
        existing_peak = ""
        if self._tracker_selected_id is not None:
            existing = next((a for a in self.applications
                             if a["id"] == self._tracker_selected_id), None)
            existing_peak = existing.get("peak_stage", "") if existing else ""

        if new_status not in TERMINAL_STATUSES:
            # new_status is a progression status — advance peak if higher
            new_order  = STAGE_ORDER.get(new_status, 0)
            peak_order = STAGE_ORDER.get(existing_peak, -1)
            peak_stage = new_status if new_order > peak_order else existing_peak
        else:
            # Terminal status — keep existing peak, or fall back to "Applied"
            peak_stage = existing_peak or "Applied"

        entry = {
            "id":             self._tracker_selected_id or datetime.now().strftime("%Y%m%d_%H%M%S%f"),
            "date_applied":   self._trk_date.get().strip(),
            "date_interview": self._trk_date_interview.get().strip(),
            "date_decision":  self._trk_date_decision.get().strip(),
            "company":        company,
            "position":       position,
            "location":       self._trk_location.get().strip(),
            "source":         self._trk_source.get().strip(),
            "status":         new_status,
            "peak_stage":     peak_stage,
            "salary_jd":      self._trk_salary_jd.get().strip(),
            "salary_req":     self._trk_salary_req.get().strip(),
            "notes":          self._trk_notes.get(1.0, "end-1c").strip(),
        }
        if self._tracker_selected_id is not None:
            for i, a in enumerate(self.applications):
                if a["id"] == self._tracker_selected_id:
                    self.applications[i] = entry
                    break
        else:
            self.applications.insert(0, entry)
            self._tracker_selected_id = entry["id"]
        self._save_applications()
        self._refresh_tracker_table()
        try:
            self._tracker_tree.selection_set(entry["id"])
            self._tracker_tree.see(entry["id"])
        except Exception:
            pass
        self._set_status(f"Saved: {company or position}")
        if hasattr(self, "_trk_status_lbl"):
            self._trk_status_lbl.configure(
                text=f"Saved  {datetime.now().strftime('%H:%M:%S')}",
                fg=C["success"])

    def _tracker_delete_entry(self):
        if not self._tracker_selected_id:
            messagebox.showwarning("No Selection", "Select an entry to delete.")
            return
        app = next((a for a in self.applications if a["id"] == self._tracker_selected_id), None)
        label = f"{app.get('company','')} — {app.get('position','')}" if app else "this entry"
        if messagebox.askyesno("Delete", f"Delete:\n{label}?"):
            self.applications = [a for a in self.applications
                                  if a["id"] != self._tracker_selected_id]
            self._save_applications()
            self._refresh_tracker_table()
            self._tracker_new_entry()
            self._set_status("Entry deleted")

    def _tracker_export_csv(self):
        if not self.applications:
            messagebox.showwarning("Empty", "No applications to export.")
            return
        fname = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"applications_{datetime.now().strftime('%Y%m%d')}.csv")
        if not fname:
            return
        fields  = ["date_applied", "date_interview", "date_decision",
                   "company", "position", "location",
                   "status", "salary_jd", "salary_req", "source", "notes"]
        headers = ["Date Applied", "Date Interview", "Date Decision",
                   "Company", "Position", "Location",
                   "Status", "Salary (JD)", "Salary Requested", "Source", "Notes"]
        try:
            with open(fname, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
                w.writerow(dict(zip(fields, headers)))
                w.writerows(self.applications)
            self._set_status(f"Exported {len(self.applications)} entries")
            messagebox.showinfo("Exported", f"Saved {len(self.applications)} entries to:\n{fname}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _tracker_import_file(self):
        """Open a CSV/XLSX file, confirm column mapping, and import into the tracker."""
        path_str = filedialog.askopenfilename(
            title="Import applications from CSV or XLSX",
            filetypes=[("CSV / XLSX files", "*.csv *.xlsx *.xls"),
                       ("CSV files", "*.csv"),
                       ("Excel files", "*.xlsx *.xls"),
                       ("All files", "*.*")])
        if not path_str:
            return
        path = Path(path_str)
        suffix = path.suffix.lower()
        try:
            if suffix == ".csv":
                headers, rows = read_csv(path)
            elif suffix in (".xlsx", ".xls"):
                headers, rows = self._read_xlsx_gui(path)
                if headers is None:   # user cancelled sheet selection
                    return
            else:
                messagebox.showerror("Unsupported File",
                                     f"Unsupported file type '{suffix}'.\nUse .csv or .xlsx")
                return
        except Exception as e:
            messagebox.showerror("Read Error", str(e))
            return

        # Remove entirely empty rows
        rows = [r for r in rows
                if any(str(v).strip() not in ("", "None", "nan")
                       for v in r if v is not None)]
        if not rows:
            messagebox.showwarning("No Data", "No data rows found in the file.")
            return

        # Auto-detect mapping, then let user confirm / adjust
        mapping = detect_mapping(headers)
        mapping = self._tracker_mapping_dialog(headers, mapping)
        if mapping is None:
            return  # user cancelled

        # Convert rows → entries
        entries = []
        for seq, row in enumerate(rows):
            entry = row_to_entry(row, mapping, seq)
            if entry["company"] or entry["position"]:
                entries.append(entry)
        if not entries:
            messagebox.showwarning("No Entries", "No valid entries found (need at least a company or position).")
            return

        # Merge / replace dialog
        if self.applications:
            choice = messagebox.askyesnocancel(
                "Merge or Replace?",
                f"applications.json already has {len(self.applications)} entries.\n\n"
                f"Yes  → Merge  (add {len(entries)} new entries on top)\n"
                f"No   → Replace  (overwrite all existing entries)\n"
                f"Cancel → Cancel import")
            if choice is None:
                return
            if choice:
                self.applications = entries + self.applications
            else:
                self.applications = entries
        else:
            self.applications = entries

        self._save_applications()
        self._refresh_tracker_table()
        self._set_status(f"Imported {len(entries)} entries from {path.name}")
        messagebox.showinfo("Import Complete",
                            f"Imported {len(entries)} entries from:\n{path.name}")

    def _read_xlsx_gui(self, path: Path):
        """Read an XLSX file; show a sheet-selection dialog if there are multiple sheets."""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Missing Dependency",
                                 "openpyxl is required for XLSX files.\n\nInstall it with:\n  pip install openpyxl")
            return None, None

        wb = openpyxl.load_workbook(path, data_only=True)

        sheet_name = wb.sheetnames[0]
        if len(wb.sheetnames) > 1:
            sheet_name = self._pick_sheet_dialog(wb.sheetnames)
            if sheet_name is None:
                return None, None   # cancelled

        ws = wb[sheet_name]
        all_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        # Find first non-empty row as header
        for i, row in enumerate(all_rows):
            if any(v is not None and str(v).strip() not in ("", "None") for v in row):
                headers = [str(v).strip() if v is not None else f"Col{j}"
                           for j, v in enumerate(row)]
                data = [r for r in all_rows[i + 1:]
                        if any(v is not None and str(v).strip() not in ("", "None") for v in r)]
                return headers, data

        return [], []

    def _pick_sheet_dialog(self, sheet_names: list[str]) -> str | None:
        """Modal dialog for selecting one of several sheets."""
        dlg = tk.Toplevel(self)
        dlg.title("Select Sheet")
        dlg.configure(bg=C["surface"])
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text="Multiple sheets found — pick one:",
                 font=(FF, 10), bg=C["surface"], fg=C["text"],
                 padx=20, pady=(14)).pack(anchor="w", padx=20, pady=(14, 6))

        var = tk.StringVar(value=sheet_names[0])
        for name in sheet_names:
            tk.Radiobutton(dlg, text=name, variable=var, value=name,
                           font=(FF, 9), bg=C["surface"], fg=C["text"],
                           selectcolor=C["surface2"], activebackground=C["surface"],
                           relief="flat").pack(anchor="w", padx=32, pady=2)

        result = [sheet_names[0]]

        def ok():
            result[0] = var.get()
            dlg.destroy()

        def cancel():
            result[0] = None
            dlg.destroy()

        btn_row = tk.Frame(dlg, bg=C["surface"])
        btn_row.pack(pady=(10, 16), padx=20, fill="x")
        self._btn(btn_row, "Select", ok, style="primary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Cancel", cancel, style="ghost").pack(side="left")

        dlg.wait_window()
        return result[0]

    def _tracker_mapping_dialog(self, headers: list[str], mapping: dict) -> dict | None:
        """Modal dialog to review and adjust auto-detected column mapping."""
        FIELD_LABELS = {
            "company":      "Company name",
            "position":     "Position / role",
            "location":     "Location",
            "status":       "Status / outcome",
            "salary_jd":    "Salary from JD",
            "salary_req":   "Salary requested",
            "date_applied": "Date applied",
            "source":       "Application source",
            "notes":        "Notes",
        }

        dlg = tk.Toplevel(self)
        dlg.title("Confirm Column Mapping")
        dlg.configure(bg=C["surface"])
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text="Auto-detected column mapping — adjust if needed:",
                 font=(FF, 10), bg=C["surface"], fg=C["text"]).grid(
            row=0, column=0, columnspan=3, padx=20, pady=(14, 10), sticky="w")

        SKIP = "— skip —"
        options = [SKIP] + headers
        vars_ = {}

        for i, (field, label) in enumerate(FIELD_LABELS.items(), start=1):
            tk.Label(dlg, text=label, font=(FF, 9), bg=C["surface"],
                     fg=C["text_muted"], anchor="e", width=20).grid(
                row=i, column=0, padx=(20, 8), pady=3, sticky="e")

            idx = mapping.get(field)
            default = headers[idx] if idx is not None else SKIP
            var = tk.StringVar(value=default)
            vars_[field] = var

            om = tk.OptionMenu(dlg, var, *options)
            om.configure(font=(FF, 9), bg=C["input_bg"], fg=C["text"],
                         relief="flat", bd=0, highlightthickness=1,
                         highlightbackground=C["border"],
                         activebackground=C["accent_dim"], width=30)
            om["menu"].configure(font=(FF, 9), bg=C["input_bg"], fg=C["text"],
                                 relief="flat", activebackground=C["accent_dim"])
            om.grid(row=i, column=1, padx=(0, 20), pady=3, sticky="ew")

        result = [None]

        def ok():
            new_mapping = {}
            for field, var in vars_.items():
                chosen = var.get()
                if chosen != SKIP and chosen in headers:
                    new_mapping[field] = headers.index(chosen)
            result[0] = new_mapping
            dlg.destroy()

        def cancel():
            dlg.destroy()

        btn_row = tk.Frame(dlg, bg=C["surface"])
        btn_row.grid(row=len(FIELD_LABELS) + 1, column=0, columnspan=2,
                     pady=(10, 16), padx=20, sticky="w")
        self._btn(btn_row, "Import", ok,     style="primary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Cancel", cancel, style="ghost").pack(side="left")

        dlg.wait_window()
        return result[0]

    def _analyze_rejection_from_history(self):
        if not self._check_init():
            return
        sel = self._hist_list.curselection()
        if not sel or sel[0] >= len(self._hist_display_indices):
            return
        e = self.history[self._hist_display_indices[sel[0]]]

        # Small dialog to paste the rejection email
        dlg = tk.Toplevel(self.root)
        dlg.title("Analyze Rejection")
        dlg.configure(bg=C["surface"])
        dlg.geometry("560x340")
        dlg.resizable(False, True)
        dlg.grab_set()

        tk.Label(dlg, text=f"Rejection: {e.get('company','')} — {e.get('role','')}",
                 font=(FF, 10, "bold"), bg=C["surface"], fg=C["text"]).pack(
            anchor="w", padx=20, pady=(16, 4))
        tk.Label(dlg, text="Paste the rejection email below (optional — leave blank to analyse from profile alone).",
                 font=(FF, 8), bg=C["surface"], fg=C["text_muted"], wraplength=520).pack(
            anchor="w", padx=20, pady=(0, 8))

        border = tk.Frame(dlg, bg=C["border_bright"], padx=1, pady=1)
        border.pack(fill="both", expand=True, padx=20)
        rejection_text = tk.Text(border, font=(FF, 9), bg=C["input_bg"], fg=C["text"],
                                 insertbackground=C["accent_soft"], relief="flat",
                                 bd=0, wrap="word", padx=8, pady=6)
        rejection_text.pack(fill="both", expand=True)
        rejection_text.bind("<FocusIn>",  lambda ev: border.configure(bg=C["accent"]))
        rejection_text.bind("<FocusOut>", lambda ev: border.configure(bg=C["border_bright"]))

        btn_row = tk.Frame(dlg, bg=C["surface"])
        btn_row.pack(fill="x", padx=20, pady=(10, 16))

        def _run():
            rejection_msg = rejection_text.get(1.0, "end").strip() or None
            dlg.destroy()
            materials = e.get("content", "") or None
            cb = self._stream_into(self.hist_content,
                                   prefix="REJECTION ANALYSIS\n" + "─" * 60 + "\n\n")
            self._set_status("Analysing rejection …")

            def _do():
                try:
                    r    = self.assistant.analyze_rejection(
                        company_name=e.get("company") or None,
                        role_title=e.get("role") or None,
                        rejection_message=rejection_msg,
                        application_materials=materials,
                        model=self.model_var.get(), stream_callback=cb)
                    toks = r.get("tokens_used", 0)
                    self._add_tokens(toks)
                    self._set_status(f"Rejection analysis complete — {toks:,} tokens")
                    self._add_history("Rejection Analysis", r.get("rejection_analysis", ""),
                                      toks, e.get("company", ""), e.get("role", ""))
                except Exception as ex:
                    self._write(self.hist_content, f"Error: {ex}")
                    self._set_status("Error analysing rejection")

            threading.Thread(target=_do, daemon=True).start()

        self._btn(btn_row, "Analyze", _run, style="primary").pack(side="left", padx=(0, 8))
        self._btn(btn_row, "Cancel", dlg.destroy, style="ghost").pack(side="left")

    def _add_to_tracker_from_history(self):
        sel = self._hist_list.curselection()
        if not sel:
            messagebox.showwarning("No Selection", "Select a history entry first.")
            return
        if sel[0] >= len(self._hist_display_indices):
            return
        h = self.history[self._hist_display_indices[sel[0]]]
        self._tracker_new_entry()
        self._trk_company.set(h.get("company", ""))
        self._trk_position.set(h.get("role",    ""))
        self._trk_date.set(datetime.now().strftime("%Y-%m-%d"))
        self._trk_status_var.set("Applied")
        note = f"Materials via assistant — {h.get('type','')} on {h.get('timestamp','')}"
        self._trk_notes.delete(1.0, "end")
        self._trk_notes.insert(1.0, note)
        self._nav_to("Tracker")
        if hasattr(self, "_trk_status_lbl"):
            self._trk_status_lbl.configure(
                text="Pre-filled from history — review and click Save Entry",
                fg=C["warning"])

    # ═════════════════════════════════════════════════════════════════════════
    # History management
    # ═════════════════════════════════════════════════════════════════════════

    def _load_history(self):
        if Path(HISTORY_FILE).exists():
            try:
                with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return []
        return []

    def _save_history(self):
        try:
            with open(HISTORY_FILE, "w", encoding="utf-8") as f:
                json.dump(self.history, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"[History] save error: {e}")

    def _add_history(self, entry_type, content, tokens=0, company=None, role=None, status=""):
        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "type":      entry_type,
            "company":   company or "",
            "role":      role    or "",
            "model":     self.model_var.get(),
            "tokens":    tokens,
            "content":   content,
            "status":    status,
        }
        # Persist the evaluation context alongside each Evaluation entry
        if entry_type == "Evaluation" and self.assistant:
            entry["evaluation_context"] = getattr(self.assistant, "_last_evaluation", "")
        self.history.insert(0, entry)
        self._save_history()
        self.root.after(0, self._refresh_hist_list)

    def _refresh_hist_list(self):
        if not hasattr(self, "_hist_list"):
            return
        search_text   = self._hist_search_var.get().strip().lower()
        filter_status = self._hist_filter_status
        status_tag = {"Applied": "[Applied]  ", "Interview": "[Intv]     ",
                      "Offer": "[Offer]    ", "Rejected": "[Rejctd]   ", "": "           "}
        self._hist_list.delete(0, "end")
        self._hist_display_indices = []
        for orig_idx, e in enumerate(self.history):
            if filter_status and e.get("status", "") != filter_status:
                continue
            if search_text:
                haystack = " ".join([e.get("company",""), e.get("role",""),
                                     e.get("type",""),    e.get("timestamp","")]).lower()
                if search_text not in haystack:
                    continue
            st  = status_tag.get(e.get("status", ""), "           ")
            tag = "[{:<12}]".format(e.get("type", ""))
            co  = "  " + e["company"] if e.get("company") else ""
            self._hist_list.insert("end", f"{e['timestamp']}  {st}{tag}{co}")
            self._hist_display_indices.append(orig_idx)
        if hasattr(self, "_hist_count_lbl"):
            shown = len(self._hist_display_indices)
            total = len(self.history)
            self._hist_count_lbl.configure(
                text=f"{shown} of {total} entries" if shown != total else f"{total} entries")

    def _on_hist_select(self, _event=None):
        sel = self._hist_list.curselection()
        if not sel or sel[0] >= len(self._hist_display_indices):
            return
        e = self.history[self._hist_display_indices[sel[0]]]
        status_txt = f" [{e.get('status')}]" if e.get("status") else ""
        meta = f"[{e['type']}]{status_txt}  {e['timestamp']}"
        if e.get("company"): meta += f"  |  {e['company']}"
        if e.get("role"):    meta += f"  —  {e['role']}"
        meta += f"  |  {e['tokens']:,} tokens  |  {e['model']}"
        if e.get("notes"):
            snippet = e["notes"][:60].replace("\n", " ")
            meta += f"  |  📝 {snippet}{'…' if len(e['notes']) > 60 else ''}"
        self._hist_meta.configure(text=meta)
        self._write(self.hist_content, e.get("content", ""))
        # Restore evaluation context so Generate/Interview pages can use it
        if self.assistant and e.get("evaluation_context"):
            self.assistant._last_evaluation = e["evaluation_context"]
            self._refresh_eval_labels()
        # Load personal notes
        if hasattr(self, "_hist_notes"):
            self._hist_notes.configure(state="normal", fg=C["text"])
            self._hist_notes.delete(1.0, "end")
            self._hist_notes.insert(1.0, e.get("notes", ""))
        # Show/hide rejection analysis button
        if hasattr(self, "_hist_rejection_btn"):
            if e.get("status") == "Rejected":
                self._hist_rejection_btn.pack(side="left")
            else:
                self._hist_rejection_btn.pack_forget()

    def _save_hist_notes(self):
        """Auto-save the notes field back to the selected history entry."""
        if not hasattr(self, "_hist_notes"):
            return
        sel = self._hist_list.curselection()
        if not sel or sel[0] >= len(self._hist_display_indices):
            return
        idx = self._hist_display_indices[sel[0]]
        notes = self._hist_notes.get(1.0, "end-1c").strip()
        self.history[idx]["notes"] = notes
        self._save_history()

    def _set_hist_status(self, status):
        sel = self._hist_list.curselection()
        if not sel:
            messagebox.showwarning("No Selection", "Select a history entry first.")
            return
        if sel[0] >= len(self._hist_display_indices):
            return
        orig_idx   = self._hist_display_indices[sel[0]]
        new_status = "" if status == "Clear" else status
        self.history[orig_idx]["status"] = new_status
        self._save_history()
        self._refresh_hist_list()
        self._set_status(f"Status set to: {new_status or 'cleared'}")

    def _delete_hist_entry(self):
        sel = self._hist_list.curselection()
        if not sel:
            messagebox.showwarning("No Selection", "Select an entry to delete.")
            return
        if sel[0] >= len(self._hist_display_indices):
            return
        if messagebox.askyesno("Delete", "Delete this history entry?"):
            del self.history[self._hist_display_indices[sel[0]]]
            self._save_history()
            self._refresh_hist_list()
            self._clear_text(self.hist_content)
            self._hist_meta.configure(text="Select an entry to view its content")

    def _set_hist_filter(self, status):
        self._hist_filter_status = status
        for s, btn in self._hist_filter_btns.items():
            active = (s == status)
            btn.configure(bg=C["accent"] if active else C["surface2"],
                          fg=C["text_inv"] if active else C["text_dim"])
        self._refresh_hist_list()

    def _clear_history(self):
        if not self.history:
            return
        if messagebox.askyesno("Clear All", "Delete all history entries?\nThis cannot be undone."):
            self.history.clear()
            self._save_history()
            self._refresh_hist_list()
            self._clear_text(self.hist_content)
            self._hist_meta.configure(text="Select an entry to view its content")

    def _history_export_csv(self):
        if not self.history:
            messagebox.showwarning("Empty", "No history entries to export.")
            return
        include_content = messagebox.askyesno(
            "Include Content?",
            "Include the full generated text in the CSV?\n\n"
            "Yes = include content column (larger file)\n"
            "No  = metadata only (timestamp, type, company, role, model, tokens, status)")
        fname = filedialog.asksaveasfilename(
            title="Export History as CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"history_{datetime.now().strftime('%Y%m%d')}.csv")
        if not fname:
            return
        fields  = ["timestamp", "type", "company", "role", "model", "tokens", "status"]
        headers = ["Timestamp",  "Type",  "Company", "Role",  "Model",  "Tokens", "Status"]
        if include_content:
            fields.append("content")
            headers.append("Content")
        try:
            with open(fname, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
                w.writerow(dict(zip(fields, headers)))
                w.writerows(self.history)
            self._set_status(f"History exported — {len(self.history)} entries")
            messagebox.showinfo("Exported",
                                f"Saved {len(self.history)} entries to:\n{fname}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _history_export_json(self):
        if not self.history:
            messagebox.showwarning("Empty", "No history entries to export.")
            return
        fname = filedialog.asksaveasfilename(
            title="Export History as JSON",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=f"history_{datetime.now().strftime('%Y%m%d')}.json")
        if not fname:
            return
        try:
            with open(fname, "w", encoding="utf-8") as f:
                json.dump(self.history, f, indent=2, ensure_ascii=False)
            self._set_status(f"History exported as JSON — {len(self.history)} entries")
            messagebox.showinfo("Exported",
                                f"Saved {len(self.history)} entries to:\n{fname}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ═════════════════════════════════════════════════════════════════════════
    # Utility helpers
    # ═════════════════════════════════════════════════════════════════════════

    def _refresh_eval_labels(self):
        if hasattr(self, "_gen_eval_lbl"):
            self._gen_eval_lbl._update()
        if hasattr(self, "_int_eval_lbl"):
            self._int_eval_lbl._update()

    def _check_init(self):
        if not self.assistant:
            messagebox.showwarning("Not Initialized",
                                   "Initialize the assistant on the Setup page first.")
            self._nav_to("Setup")
            return False
        return True

    def _validate_jd(self, jd: str) -> bool:
        """Return True if the job description looks usable; warn and return False otherwise."""
        if not jd.strip():
            messagebox.showwarning("Input Required", "Please paste a job description.")
            return False
        if len(jd.strip()) < 100:
            return messagebox.askyesno(
                "Short Input",
                f"The job description is very short ({len(jd.strip())} characters).\n\n"
                "This may produce poor results. Continue anyway?")
        # Check CV extraction quality once (show once per session)
        if self.assistant and len(self.assistant.cv_text.strip()) < 500:
            if not getattr(self, "_cv_warned", False):
                self._cv_warned = True
                messagebox.showwarning(
                    "CV Extraction Warning",
                    "The extracted CV text is very short — it may not have parsed correctly.\n\n"
                    "Check the CV on the Setup page (Preview CV Text) before generating materials.")
        return True

    def _toggle_theme(self):
        """Switch between dark and light themes by updating C in-place and recolouring all widgets."""
        is_dark = C["bg"] == C_DARK["bg"]
        new_palette = C_LIGHT if is_dark else C_DARK
        C.update(new_palette)
        self._theme_btn.configure(text="☀" if not is_dark else "🌙")

        def _recolour(widget):
            try:
                cls = widget.winfo_class()
                if cls in ("Frame",):
                    bg = widget.cget("bg")
                    # Map old colour to new if it's a known palette colour
                    for key, old_val in (C_DARK if is_dark else C_LIGHT).items():
                        if bg == old_val and key in C:
                            widget.configure(bg=C[key])
                            break
                elif cls in ("Label", "Button"):
                    for attr in ("bg", "fg", "activebackground", "activeforeground"):
                        try:
                            val = widget.cget(attr)
                            src = C_DARK if is_dark else C_LIGHT
                            for key, old_val in src.items():
                                if val == old_val and key in C:
                                    widget.configure(**{attr: C[key]})
                                    break
                        except Exception:
                            pass
                elif cls == "Text":
                    for attr in ("bg", "fg", "insertbackground"):
                        try:
                            val = widget.cget(attr)
                            src = C_DARK if is_dark else C_LIGHT
                            for key, old_val in src.items():
                                if val == old_val and key in C:
                                    widget.configure(**{attr: C[key]})
                                    break
                        except Exception:
                            pass
                elif cls == "Entry":
                    for attr in ("bg", "fg", "insertbackground"):
                        try:
                            val = widget.cget(attr)
                            src = C_DARK if is_dark else C_LIGHT
                            for key, old_val in src.items():
                                if val == old_val and key in C:
                                    widget.configure(**{attr: C[key]})
                                    break
                        except Exception:
                            pass
            except Exception:
                pass
            for child in widget.winfo_children():
                _recolour(child)

        _recolour(self.root)
        # Persist preference
        self._save_config()

    def _adjust_font_size(self, delta: int):
        """Increase or decrease the font size of all output Text widgets."""
        new_size = max(8, min(16, self._font_size + delta))
        if new_size == self._font_size and delta != 0:
            return
        self._font_size = new_size
        # Update all Text widgets in the content area
        def _apply(widget):
            if isinstance(widget, tk.Text):
                try:
                    current = widget.cget("font")
                    # Only resize output/input areas, not tiny status labels
                    fam = FM if "Consolas" in str(current) else FF
                    widget.configure(font=(fam, new_size))
                except Exception:
                    pass
            for child in widget.winfo_children():
                _apply(child)
        _apply(self.content)
        self._save_config()

    def _set_status(self, msg):
        self.root.after(0, lambda: self._status_lbl.configure(text=msg))

    def _add_tokens(self, n):
        self._session_tokens += n
        if self.provider == "ollama":
            # Local inference — no cost
            cost_str = "Local (free)"
        else:
            rate = MODEL_COSTS.get(self.model_var.get(), 0.010)
            self._session_cost += (n / 1000) * rate
            cost_str = f"${self._session_cost:.3f}"
        toks = self._session_tokens
        self.root.after(0, lambda: self._token_lbl.configure(
            text=f"Session: {toks:,} tokens  |  {cost_str}"))

    def _write(self, widget, text):
        def _do():
            widget.configure(state="normal")
            widget.delete(1.0, "end")
            widget.insert("end", text)
            widget.configure(state="disabled")
            self._update_count(widget)
        self.root.after(0, _do)

    def _paste_to(self, widget):
        try:
            text = self.root.clipboard_get()
            widget.delete(1.0, "end")
            widget.insert(1.0, text)
        except Exception:
            messagebox.showwarning("Clipboard", "Nothing to paste.")

    def _import_jd_from_url(self, target_widget):
        """Show a URL input dialog, fetch the page, extract text, paste into target_widget."""
        from tkinter import simpledialog
        url = simpledialog.askstring(
            "Import from URL",
            "Paste the job posting URL:\n(LinkedIn, Greenhouse, Workday, Lever, etc.)",
            parent=self.root)
        if not url:
            return
        url = url.strip()
        if not url.startswith(("http://", "https://")):
            url = "https://" + url

        self._set_status("Fetching URL …")

        def _fetch():
            text, error = self._fetch_url_text(url)
            def _apply():
                if error:
                    messagebox.showerror("Import Failed", error)
                    self._set_status("URL import failed")
                    return
                if len(text) < 200:
                    messagebox.showwarning(
                        "Short Content",
                        "The fetched page contains very little text.\n\n"
                        "This may be a JavaScript-rendered page (LinkedIn, Workday).\n"
                        "Try copying and pasting the job description manually.")
                    if not text.strip():
                        self._set_status("URL import: no content extracted")
                        return
                target_widget.delete(1.0, "end")
                target_widget.insert(1.0, text)
                self._set_status(f"Imported {len(text):,} characters from URL")
            self.root.after(0, _apply)

        threading.Thread(target=_fetch, daemon=True).start()

    def _fetch_url_text(self, url: str) -> tuple[str, str]:
        """Fetch a URL and return (extracted_text, error_message). One of them will be empty."""
        try:
            import requests
        except ImportError:
            return "", ("requests is not installed.\n\nInstall it with:\n  pip install requests beautifulsoup4")
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            return "", ("beautifulsoup4 is not installed.\n\nInstall it with:\n  pip install requests beautifulsoup4")

        try:
            headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                )
            }
            response = requests.get(url, headers=headers, timeout=12)
            response.raise_for_status()
        except Exception as e:
            return "", f"Could not fetch URL:\n{e}"

        soup = BeautifulSoup(response.text, "html.parser")

        # Remove noise tags
        for tag in soup(["script", "style", "nav", "header", "footer",
                          "aside", "noscript", "iframe", "svg"]):
            tag.decompose()

        # Try to find the main content block first
        main = (soup.find("main") or soup.find("article") or
                soup.find(attrs={"role": "main"}) or
                soup.find("div", class_=re.compile(
                    r"job[-_]?(description|details|content|post|listing)", re.I)) or
                soup.find("div", id=re.compile(
                    r"job[-_]?(description|details|content|post|listing)", re.I)))

        target = main if main else soup.body if soup.body else soup

        # Extract and clean text
        lines = []
        for line in target.get_text(separator="\n").splitlines():
            stripped = line.strip()
            if stripped:
                lines.append(stripped)

        text = "\n".join(lines)
        # Collapse runs of 3+ blank lines to 2
        text = re.sub(r'\n{3,}', '\n\n', text)
        return text.strip(), ""

    def _save_text(self, widget):
        content = widget.get(1.0, "end").strip()
        if not content:
            messagebox.showwarning("Empty", "Nothing to save.")
            return
        fname = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("Markdown", "*.md")],
            initialfile=f"application_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
        if fname:
            try:
                with open(fname, "w", encoding="utf-8") as f:
                    f.write(content)
                self._set_status(f"Saved: {Path(fname).name}")
            except Exception as e:
                messagebox.showerror("Save Error", str(e))

    def _copy_text(self, widget):
        content = widget.get(1.0, "end").strip()
        if not content:
            messagebox.showwarning("Empty", "Nothing to copy.")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(content)
        self._set_status("Copied to clipboard")

    def _clear_text(self, widget):
        widget.configure(state="normal")
        widget.delete(1.0, "end")
        widget.configure(state="disabled")
        self._update_count(widget)


# ─────────────────────────────────────────────────────────────────────────────
def main():
    root = tk.Tk()
    app = JobAssistantV3(root)
    root.protocol("WM_DELETE_WINDOW", lambda: (app._save_config(), root.destroy()))
    root.mainloop()


if __name__ == "__main__":
    main()
